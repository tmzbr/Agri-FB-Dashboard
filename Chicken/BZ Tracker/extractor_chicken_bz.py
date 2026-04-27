#!/usr/bin/env python3
"""
extractor_chicken_bz.py — Brazil Chicken Export Spread Tracker
===============================================================
Builds / refreshes  chicken_bz.db  with monthly and weekly spread data.

DATA SOURCES
  • SECEX/MDIC monthly exports  → balanca.economia.gov.br annual CSVs
                                   NCM 0207 (poultry — all sub-codes)
  • BCB PTAX BRL/USD daily FX   → BCB OLINDA API (primary) / BCB SGS (fallback)
  • CEPEA grain costs (corn+soy)→ CEPEA_MONTHLY_SEED dict embedded in this file
                                   (corn Paranaguá + soy PNA, BRL/sc60kg monthly avgs)
                                   Grain basket: 66% corn + 34% soy (by sc60kg weight)
                                   Applied with 2-month lag to spread computation.
                                   → Update CEPEA_MONTHLY_SEED every ~6 months by
                                   appending new "YYYY-MM": (corn_avg, soy_avg) entries.
                                   Optionally: pass --cepea PATH to seed from xlsx file.

USAGE
  pip install requests
  python extractor_chicken_bz.py                        # incremental (SECEX + FX + CEPEA seed)
  python extractor_chicken_bz.py --init                 # full reseed from scratch
  python extractor_chicken_bz.py --cepea PATH           # also import CEPEA from xlsx file

OUTPUT
  chicken_bz.db  (SQLite)

SCHEMA — table: monthly
  period       TEXT PRIMARY KEY   e.g. "2022-08"
  year         INTEGER
  month        INTEGER
  secex_usd_kg REAL               FOB USD / kg net weight
  fx           REAL               avg BCB PTAX BRL/USD for the month
  secex_brl_kg REAL               secex_usd_kg × fx
  cepea_r_kg   REAL               grain basket BRL/kg (corn66+soy34, 2-mo lag)
  spread       REAL               (secex_brl_kg - grain_brl_kg) / secex_brl_kg
  updated_at   TEXT               ISO timestamp

SCHEMA — table: weekly
  start_date   TEXT PRIMARY KEY   ISO date "YYYY-MM-DD"
  end_date     TEXT               ISO date "YYYY-MM-DD"
  secex_usd_kg REAL               from SECEX weekly cumulative data
  fx           REAL               avg BCB PTAX BRL/USD for the period
  secex_brl_kg REAL               secex_usd_kg × fx
  cepea_r_kg   REAL               grain basket BRL/kg (corn66+soy34, 2-mo lag)
  spread       REAL               (secex_brl_kg - grain_brl_kg) / secex_brl_kg
  vol_tons     REAL               incremental weekly export tons
  updated_at   TEXT               ISO timestamp
"""

import sqlite3, os, sys, time, warnings
from datetime import datetime, date
from calendar import monthrange
from pathlib import Path

try:
    import requests
    from urllib3.exceptions import InsecureRequestWarning
    warnings.filterwarnings("ignore", category=InsecureRequestWarning)
except ImportError:
    sys.exit("Missing: pip install requests")

_NO_VERIFY_HOSTS = ("balanca.economia.gov.br", "olinda.bcb.gov.br", "api.bcb.gov.br")

# ── Brazil business-day helpers ───────────────────────────────────────────────
def _easter(year):
    a = year % 19; b = year // 100; c = year % 100
    d = b // 4;    e = b % 4;       f = (b + 8) // 25
    g = (b - f + 1) // 3;  h = (19*a + b - d - g + 15) % 30
    i = c // 4;    k = c % 4;       l = (32 + 2*e + 2*i - h - k) % 7
    m = (a + 11*h + 22*l) // 451
    mo  = (h + l - 7*m + 114) // 31
    day = ((h + l - 7*m + 114) % 31) + 1
    return date(year, mo, day)


def _br_holidays(year):
    from datetime import timedelta
    e = _easter(year)
    h = {
        str(date(year, 1,  1)),
        str(e - timedelta(days=48)),
        str(e - timedelta(days=47)),
        str(e - timedelta(days=2)),
        str(date(year, 4,  21)),
        str(date(year, 5,  1)),
        str(date(year, 9,  7)),
        str(date(year, 10, 12)),
        str(date(year, 11, 2)),
        str(date(year, 11, 15)),
        str(date(year, 12, 25)),
    }
    if year >= 2024:
        h.add(str(date(year, 11, 20)))
    return h


def _biz_days_between(start_dt, end_dt):
    """Count Mon–Fri days (excl. BR national holidays) between dates, inclusive."""
    from datetime import timedelta
    hols = _br_holidays(start_dt.year)
    if end_dt.year != start_dt.year:
        hols |= _br_holidays(end_dt.year)
    count = 0
    d = start_dt
    while d <= end_dt:
        if d.weekday() < 5 and str(d) not in hols:
            count += 1
        d += timedelta(days=1)
    return count


# ── Paths ──────────────────────────────────────────────────────────────────────
DB_PATH   = Path(__file__).parent / "chicken_bz.db"
TIMEOUT   = 30
RETRY     = 3
NCM_CODE  = "0207"    # 4-digit prefix for all poultry NCM codes
ANO_INI   = 2004      # earliest SECEX year to seed

# Grain basket weights and lag
CORN_WEIGHT = 0.66
SOY_WEIGHT  = 0.34
GRAIN_LAG   = 2  # months

# ── Weekly historical data (Aug 2022 – Apr 2026, 177 weeks) ───────────────────
# vol_tons = MTD cumulative tons (materialise() de-accumulates to weekly incremental)
WEEKLY_SEED = [
    # (start_date, end_date, price_usd_kg, vol_tons_mtd_cumulative)
    # ── 2022 ────────────────────────────────────────────────────────────────────
    ('2022-08-01', '2022-08-07', 2.1294, 105352.4),
    ('2022-08-08', '2022-08-14', 2.1037, 192706.6),
    ('2022-08-15', '2022-08-21', 2.0421, 254021.3),
    ('2022-08-22', '2022-08-31', 2.0574, 398599.3),
    ('2022-09-01', '2022-09-11', 2.0203, 131867.8),
    ('2022-09-12', '2022-09-18', 2.1256, 219898.5),
    ('2022-09-19', '2022-09-25', 2.1306, 294708.2),
    ('2022-09-26', '2022-10-02', 2.0288, 364238.4),
    ('2022-10-03', '2022-10-09', 2.0567, 114110.1),
    ('2022-10-10', '2022-10-16', 2.0249, 201664.0),
    ('2022-10-17', '2022-10-23', 2.0759, 287533.6),
    ('2022-10-24', '2022-10-31', 2.1357, 362940.2),
    ('2022-11-01', '2022-11-13', 2.0505, 161732.5),
    ('2022-11-14', '2022-11-20', 2.172,  220208.8),
    ('2022-11-21', '2022-11-30', 2.0504, 346707.9),
    ('2022-12-01', '2022-12-11', 2.0127, 108923.5),
    ('2022-12-12', '2022-12-18', 1.9758, 193378.8),
    ('2022-12-19', '2022-12-25', 2.0216, 283161.4),
    ('2022-12-26', '2023-01-01', 2.0134, 353629.4),
    # ── 2023 ────────────────────────────────────────────────────────────────────
    ('2023-01-02', '2023-01-08', 2.0248, 117227.5),
    ('2023-01-09', '2023-01-15', 1.8751, 193701.7),
    ('2023-01-16', '2023-01-22', 2.0379, 270582.6),
    ('2023-01-23', '2023-01-31', 2.011,  388597.6),
    ('2023-02-01', '2023-02-12', 1.8548, 140338.7),
    ('2023-02-13', '2023-02-19', 1.8836, 247786.6),
    ('2023-02-20', '2023-02-28', 1.935,  353421.9),
    ('2023-03-01', '2023-03-12', 1.8928, 189171.0),
    ('2023-03-13', '2023-03-19', 1.8626, 291347.2),
    ('2023-03-20', '2023-04-02', 1.8544, 397163.1),
    ('2023-04-03', '2023-04-09', 1.9183, 130522.1),
    ('2023-04-10', '2023-04-16', 1.82,   229508.1),
    ('2023-04-17', '2023-04-23', 1.9,    310377.6),
    ('2023-04-24', '2023-04-30', 1.96,   408278.2),
    ('2023-05-01', '2023-05-07', 1.9874, 103242.4),
    ('2023-05-08', '2023-05-14', 1.9594, 192790.3),
    ('2023-05-15', '2023-05-21', 1.9424, 275919.9),
    ('2023-05-22', '2023-05-31', 1.9302, 402769.3),
    ('2023-06-01', '2023-06-11', 1.984,  162407.5),
    ('2023-06-12', '2023-06-18', 2.067,  249799.4),
    ('2023-06-19', '2023-06-25', 1.9576, 344201.5),
    ('2023-06-26', '2023-07-02', 1.9363, 419253.6),
    ('2023-07-03', '2023-07-09', 1.9364, 98913.7),
    ('2023-07-10', '2023-07-16', 2.004,  214648.4),
    ('2023-07-17', '2023-07-23', 1.9331, 298493.0),
    ('2023-07-24', '2023-07-31', 1.8998, 404605.6),
    ('2023-08-01', '2023-08-06', 1.6907, 117625.7),
    ('2023-08-07', '2023-08-13', 1.8826, 211487.5),
    ('2023-08-14', '2023-08-20', 1.8968, 287635.2),
    ('2023-08-21', '2023-08-31', 1.8195, 416380.2),
    ('2023-09-01', '2023-09-10', 1.7741, 147968.0),
    ('2023-09-11', '2023-09-17', 1.7916, 240753.4),
    ('2023-09-18', '2023-09-24', 1.7835, 318003.9),
    ('2023-09-25', '2023-10-01', 1.7425, 373310.6),
    ('2023-10-02', '2023-10-08', 1.7766, 102085.3),
    ('2023-10-09', '2023-10-15', 1.8112, 187963.4),
    ('2023-10-16', '2023-10-22', 1.7823, 252382.1),
    ('2023-10-23', '2023-10-31', 1.7228, 374536.6),
    ('2023-11-01', '2023-11-12', 1.7453, 148821.9),
    ('2023-11-13', '2023-11-19', 1.7581, 224887.4),
    ('2023-11-20', '2023-11-30', 1.8155, 356342.7),
    ('2023-12-01', '2023-12-10', 1.7534, 151555.2),
    ('2023-12-11', '2023-12-17', 1.7625, 244059.7),
    ('2023-12-18', '2023-12-24', 1.6699, 346449.2),
    ('2023-12-25', '2023-12-31', 1.6731, 435289.2),
    # ── 2024 ────────────────────────────────────────────────────────────────────
    ('2024-01-01', '2024-01-07', 1.6138, 105082.1),
    ('2024-01-08', '2024-01-14', 1.6872, 170530.8),
    ('2024-01-15', '2024-01-21', 1.6001, 233126.6),
    ('2024-01-22', '2024-01-28', 1.6852, 332350.6),
    ('2024-01-29', '2024-01-31', 1.6325, 375949.9),
    ('2024-02-01', '2024-02-11', 1.7342, 141690.8),
    ('2024-02-12', '2024-02-25', 1.7057, 311039.9),
    ('2024-02-26', '2024-02-29', 1.8183, 368664.9),
    ('2024-03-01', '2024-03-10', 1.808,  145742.1),
    ('2024-03-11', '2024-03-17', 1.7181, 226608.4),
    ('2024-03-18', '2024-03-24', 1.8108, 314097.4),
    ('2024-03-25', '2024-03-31', 1.6561, 391278.6),
    ('2024-04-01', '2024-04-07', 1.8222, 107393.4),
    ('2024-04-08', '2024-04-14', 1.8311, 197901.2),
    ('2024-04-15', '2024-04-21', 1.7997, 309523.3),
    ('2024-04-22', '2024-04-28', 1.7684, 421579.3),
    ('2024-04-29', '2024-04-30', 1.7999, 453012.7),
    ('2024-05-01', '2024-05-12', 1.7452, 169394.1),
    ('2024-05-13', '2024-06-02', 1.7886, 424917.9),
    ('2024-06-03', '2024-06-09', 1.7261, 118873.5),
    ('2024-06-10', '2024-06-16', 1.8442, 210538.1),
    ('2024-06-17', '2024-06-23', 1.739,  313419.6),
    ('2024-06-24', '2024-06-30', 1.8513, 408543.1),
    ('2024-07-01', '2024-07-07', 1.8368, 121683.0),
    ('2024-07-08', '2024-07-14', 1.9412, 223500.0),
    ('2024-07-15', '2024-07-21', 1.9326, 301651.2),
    ('2024-07-22', '2024-07-28', 1.8998, 397829.7),
    ('2024-07-29', '2024-07-31', 1.818,  436699.5),
    ('2024-08-01', '2024-08-11', 1.7826, 145164.8),
    ('2024-08-12', '2024-08-18', 1.7901, 194252.8),
    ('2024-08-19', '2024-08-25', 2.0283, 281526.7),
    ('2024-08-26', '2024-09-01', 2.8655, 356445.3),
    ('2024-09-02', '2024-09-08', 1.8863, 116660.1),
    ('2024-09-09', '2024-09-15', 1.9201, 241719.2),
    ('2024-09-16', '2024-09-22', 1.9219, 318868.1),
    ('2024-09-23', '2024-09-30', 1.9457, 451796.4),
    ('2024-10-01', '2024-10-06', 1.9584, 80901.5),
    ('2024-10-07', '2024-10-13', 1.9466, 180943.7),
    ('2024-10-14', '2024-10-20', 1.9077, 286609.8),
    ('2024-10-21', '2024-10-27', 1.8532, 380961.0),
    ('2024-10-28', '2024-11-03', 1.8381, 434767.5),
    ('2024-11-04', '2024-11-10', 1.9134, 174340.6),
    ('2024-11-11', '2024-11-17', 1.8642, 267348.7),
    ('2024-11-18', '2024-11-24', 1.8352, 353675.9),
    ('2024-11-25', '2024-12-01', 1.8506, 436652.4),
    ('2024-12-02', '2024-12-08', 1.8735, 118371.1),
    ('2024-12-09', '2024-12-15', 1.8466, 209641.2),
    ('2024-12-16', '2024-12-22', 1.8489, 288483.3),
    ('2024-12-23', '2025-01-01', 1.8262, 413456.2),
    # ── 2025 ────────────────────────────────────────────────────────────────────
    ('2025-01-02', '2025-01-12', 1.7987, 164509.7),
    ('2025-01-13', '2025-01-19', 1.7825, 263486.7),
    ('2025-01-20', '2025-01-26', 1.7437, 349327.4),
    ('2025-01-27', '2025-02-02', 1.9885, 415533.7),
    ('2025-02-03', '2025-02-09', 1.7931, 143443.8),
    ('2025-02-10', '2025-02-16', 1.7482, 259780.3),
    ('2025-02-17', '2025-02-23', 1.7932, 355938.0),
    ('2025-02-24', '2025-03-02', 1.8178, 436752.4),
    ('2025-03-03', '2025-03-09', 1.8063, 131247.5),
    ('2025-03-10', '2025-03-16', 1.7558, 227768.1),
    ('2025-03-17', '2025-03-23', 1.7663, 323795.9),
    ('2025-03-24', '2025-03-31', 1.8311, 438916.7),
    ('2025-04-01', '2025-04-06', 1.7884, 87245.3),
    ('2025-04-07', '2025-04-13', 1.8064, 201349.2),
    ('2025-04-14', '2025-04-20', 1.8732, 283375.6),
    ('2025-04-21', '2025-04-27', 1.8436, 389401.3),
    ('2025-04-28', '2025-05-04', 1.8991, 440665.9),
    ('2025-05-05', '2025-05-11', 1.8008, 139401.4),
    ('2025-05-12', '2025-05-18', 1.828,  222858.9),
    ('2025-05-19', '2025-05-25', 1.8015, 318529.3),
    ('2025-05-26', '2025-06-01', 1.7656, 363108.3),
    ('2025-06-02', '2025-06-08', 1.7917, 89765.8),
    ('2025-06-09', '2025-06-15', 1.8083, 152448.3),
    ('2025-06-16', '2025-06-22', 1.7802, 224039.9),
    ('2025-06-23', '2025-06-30', 1.811,  313807.9),
    ('2025-07-01', '2025-07-06', 1.7487, 75157.2),
    ('2025-07-07', '2025-07-13', 1.8038, 150289.6),
    ('2025-07-14', '2025-07-20', 1.8683, 231985.2),
    ('2025-07-21', '2025-07-27', 1.829,  329030.7),
    ('2025-07-28', '2025-08-03', 1.8342, 375982.7),
    ('2025-08-04', '2025-08-10', 1.8004, 112875.9),
    ('2025-08-11', '2025-08-17', 1.7374, 201826.8),
    ('2025-08-18', '2025-08-24', 1.7816, 283938.9),
    ('2025-08-25', '2025-08-31', 1.6832, 373989.6),
    ('2025-09-01', '2025-09-07', 1.7692, 125826.9),
    ('2025-09-08', '2025-09-14', 1.773,  234428.6),
    ('2025-09-15', '2025-09-21', 1.75,   333265.6),
    ('2025-09-22', '2025-09-28', 1.7637, 440502.3),
    ('2025-09-29', '2025-10-05', 1.8418, 459823.3),
    ('2025-10-06', '2025-10-12', 1.6814, 183288.7),
    ('2025-10-13', '2025-10-19', 1.3291, 315809.7),
    ('2025-10-20', '2025-10-26', 2.2702, 395074.7),
    ('2025-10-27', '2025-11-02', 1.6456, 474017.2),
    ('2025-11-03', '2025-11-09', 1.8113, 135114.9),
    ('2025-11-10', '2025-11-16', 1.6821, 214595.3),
    ('2025-11-17', '2025-11-23', 1.3344, 323679.0),
    ('2025-11-24', '2025-11-30', 2.4202, 400896.7),
    ('2025-12-01', '2025-12-07', 1.6874, 128092.2),
    ('2025-12-08', '2025-12-14', 1.7817, 229157.0),
    ('2025-12-15', '2025-12-21', 1.7497, 327390.4),
    ('2025-12-22', '2026-01-04', 1.8038, 469906.1),
    # ── 2026 ────────────────────────────────────────────────────────────────────
    ('2026-01-05', '2026-01-11', 1.8104, 181046.4),
    ('2026-01-12', '2026-01-18', 1.7614, 243799.9),
    ('2026-01-19', '2026-01-25', 1.7838, 349691.9),
    ('2026-01-26', '2026-02-01', 2.0826, 430379.9),
    ('2026-02-02', '2026-02-08', 1.8494, 142801.1),
    ('2026-02-09', '2026-02-15', 1.8302, 263918.2),
    ('2026-02-16', '2026-02-22', 1.8735, 376560.2),
    ('2026-02-23', '2026-03-01', 1.8928, 460606.2),
    ('2026-03-02', '2026-03-08', 1.7895, 132314.7),
    ('2026-03-09', '2026-03-15', 1.8723, 226759.6),
    ('2026-03-16', '2026-03-22', 1.8362, 329818.3),
    ('2026-03-23', '2026-03-31', 1.8246, 468706.4),
    ('2026-04-01', '2026-04-10', 1.8543, 183691.4),
]

# ── CEPEA grain monthly averages (corn Paranaguá + soy PNA, BRL/sc60kg) ───────
# Generated from CEPEA_Daily historical data. Covers 2006-03 → 2026-04.
# To extend: append new entries at the end — e.g.:
#   "2026-05": (71.2, 130.5),
# Values are monthly simple averages of daily CEPEA/ESALQ closing prices.
CEPEA_MONTHLY_SEED = {
    "2006-03": (14.01, 27.79),  "2006-04": (14.44, 27.02),  "2006-05": (15.25, 28.68),
    "2006-06": (16.49, 29.41),  "2006-07": (16.69, 29.6),   "2006-08": (16.86, 28.91),
    "2006-09": (17.94, 29.1),   "2006-10": (21.09, 32.13),  "2006-11": (22.93, 34.37),
    "2006-12": (24.96, 33.31),
    "2007-01": (25.02, 33.68),  "2007-02": (22.02, 34.71),  "2007-03": (20.2,  34.01),
    "2007-04": (19.21, 31.96),  "2007-05": (18.93, 32.12),  "2007-06": (19.58, 32.94),
    "2007-07": (18.97, 33.63),  "2007-08": (22.13, 37.11),  "2007-09": (26.95, 41.3),
    "2007-10": (27.36, 42.2),   "2007-11": (31.72, 42.96),  "2007-12": (33.8,  44.03),
    "2008-01": (30.93, 47.93),  "2008-02": (27.79, 49.99),  "2008-03": (27.19, 48.06),
    "2008-04": (26.62, 46.72),  "2008-05": (27.43, 46.65),  "2008-06": (26.88, 52.35),
    "2008-07": (27.76, 52.88),  "2008-08": (24.56, 46.84),  "2008-09": (23.78, 48.14),
    "2008-10": (22.32, 46.65),  "2008-11": (20.51, 47.32),  "2008-12": (20.75, 46.73),
    "2009-01": (23.67, 51.02),  "2009-02": (22.26, 49.24),  "2009-03": (20.62, 47.67),
    "2009-04": (21.29, 50.16),  "2009-05": (22.25, 52.78),  "2009-06": (22.24, 52.37),
    "2009-07": (20.55, 49.94),  "2009-08": (19.42, 50.06),  "2009-09": (19.12, 46.27),
    "2009-10": (20.6,  43.96),  "2009-11": (20.41, 42.01),  "2009-12": (20.02, 41.39),
    "2010-01": (19.66, 40.22),  "2010-02": (18.35, 38.24),  "2010-03": (18.47, 37.38),
    "2010-04": (18.16, 37.43),  "2010-05": (18.67, 38.4),   "2010-06": (19.43, 38.91),
    "2010-07": (18.84, 41.4),   "2010-08": (20.56, 43.83),  "2010-09": (24.36, 44.75),
    "2010-10": (25.15, 45.72),  "2010-11": (28.29, 49.17),  "2010-12": (28.36, 49.59),
    "2011-01": (30.35, 50.78),  "2011-02": (31.68, 51.39),  "2011-03": (31.44, 49.54),
    "2011-04": (29.94, 47.19),  "2011-05": (28.69, 47.83),  "2011-06": (30.75, 47.88),
    "2011-07": (30.31, 48.5),   "2011-08": (30.2,  49.38),  "2011-09": (31.92, 51.94),
    "2011-10": (30.75, 48.47),  "2011-11": (29.81, 47.74),  "2011-12": (28.18, 47.7),
    "2012-01": (31.08, 49.55),  "2012-02": (28.4,  49.32),  "2012-03": (28.89, 54.93),
    "2012-04": (25.83, 60.35),  "2012-05": (24.91, 63.79),  "2012-06": (24.13, 68.05),
    "2012-07": (29.01, 79.72),  "2012-08": (33.25, 85.58),  "2012-09": (32.23, 86.82),
    "2012-10": (31.35, 75.73),  "2012-11": (34.09, 75.73),  "2012-12": (34.96, 75.73),
    "2013-01": (32.75, 71.99),  "2013-02": (32.34, 64.01),  "2013-03": (30.71, 61.84),
    "2013-04": (26.41, 59.45),  "2013-05": (26.02, 61.89),  "2013-06": (26.45, 68.72),
    "2013-07": (25.0,  69.29),  "2013-08": (24.04, 69.88),  "2013-09": (25.07, 73.47),
    "2013-10": (24.12, 73.84),  "2013-11": (25.59, 76.35),  "2013-12": (26.45, 77.25),
    "2014-01": (26.83, 72.29),  "2014-02": (30.62, 69.71),  "2014-03": (32.84, 72.27),
    "2014-04": (31.18, 71.11),  "2014-05": (28.75, 70.74),  "2014-06": (26.38, 70.86),
    "2014-07": (23.66, 67.2),   "2014-08": (22.91, 67.11),  "2014-09": (22.02, 63.06),
    "2014-10": (23.62, 61.17),  "2014-11": (27.66, 61.17),  "2014-12": (27.67, 61.17),
    "2015-01": (27.41, 61.14),  "2015-02": (27.99, 63.72),  "2015-03": (29.44, 67.9),
    "2015-04": (27.61, 69.53),  "2015-05": (25.34, 66.61),  "2015-06": (25.03, 67.88),
    "2015-07": (25.99, 72.89),  "2015-08": (27.4,  77.33),  "2015-09": (31.04, 81.35),
    "2015-10": (32.83, 81.98),  "2015-11": (33.57, 79.97),  "2015-12": (35.33, 80.76),
    "2016-01": (41.65, 82.75),  "2016-02": (42.98, 77.83),  "2016-03": (47.79, 74.53),
    "2016-04": (48.92, 78.04),  "2016-05": (51.48, 86.43),  "2016-06": (49.12, 95.19),
    "2016-07": (44.42, 87.46),  "2016-08": (45.43, 81.69),  "2016-09": (41.91, 79.5),
    "2016-10": (42.12, 76.7),   "2016-11": (38.77, 78.27),  "2016-12": (38.29, 78.43),
    "2017-01": (35.92, 76.03),  "2017-02": (36.21, 73.86),  "2017-03": (33.77, 70.01),
    "2017-04": (28.32, 65.82),  "2017-05": (27.76, 68.94),  "2017-06": (26.75, 68.95),
    "2017-07": (26.33, 72.24),  "2017-08": (26.67, 69.83),  "2017-09": (29.11, 70.41),
    "2017-10": (31.26, 71.47),  "2017-11": (31.75, 73.87),  "2017-12": (32.38, 74.24),
    "2018-01": (32.7,  71.83),  "2018-02": (34.76, 74.72),  "2018-03": (41.37, 79.39),
    "2018-04": (39.92, 85.53),  "2018-05": (42.69, 86.12),  "2018-06": (40.55, 84.83),
    "2018-07": (37.22, 88.29),  "2018-08": (41.17, 89.91),  "2018-09": (40.31, 95.48),
    "2018-10": (36.43, 90.53),  "2018-11": (36.56, 84.16),  "2018-12": (37.83, 81.1),
    "2019-01": (38.91, 76.89),  "2019-02": (40.89, 77.73),  "2019-03": (39.82, 78.27),
    "2019-04": (36.42, 76.56),  "2019-05": (34.84, 78.36),  "2019-06": (38.04, 81.9),
    "2019-07": (37.1,  78.82),  "2019-08": (36.41, 85.08),  "2019-09": (37.64, 86.5),
    "2019-10": (41.51, 88.25),  "2019-11": (44.54, 89.87),  "2019-12": (48.16, 88.15),
    "2020-01": (51.07, 87.39),  "2020-02": (51.69, 87.61),  "2020-03": (57.41, 94.97),
    "2020-04": (52.92, 102.3),  "2020-05": (50.12, 110.41), "2020-06": (47.76, 109.76),
    "2020-07": (49.7,  116.05), "2020-08": (56.62, 128.59), "2020-09": (60.06, 141.2),
    "2020-10": (72.71, 159.64), "2020-11": (80.31, 164.99), "2020-12": (75.33, 152.56),
    "2021-01": (83.65, 167.87), "2021-02": (83.89, 166.38), "2021-03": (91.51, 171.87),
    "2021-04": (97.15, 177.1),  "2021-05": (100.72,176.39), "2021-06": (92.09, 162.08),
    "2021-07": (97.48, 167.6),  "2021-08": (98.64, 171.06), "2021-09": (92.44, 172.73),
    "2021-10": (89.92, 171.17), "2021-11": (84.19, 165.79), "2021-12": (88.03, 170.25),
    "2022-01": (96.04, 179.67), "2022-02": (96.85, 195.02), "2022-03": (99.69, 199.6),
    "2022-04": (88.78, 186.36), "2022-05": (87.36, 193.38), "2022-06": (85.64, 194.97),
    "2022-07": (81.98, 190.74), "2022-08": (82.52, 187.18), "2022-09": (84.06, 187.26),
    "2022-10": (84.53, 183.73), "2022-11": (84.99, 186.13), "2022-12": (86.01, 182.05),
    "2023-01": (86.11, 177.03), "2023-02": (85.74, 172.61), "2023-03": (84.88, 162.12),
    "2023-04": (74.85, 145.24), "2023-05": (58.16, 138.11), "2023-06": (55.04, 136.45),
    "2023-07": (54.98, 146.84), "2023-08": (53.34, 148.55), "2023-09": (54.63, 147.19),
    "2023-10": (59.13, 144.09), "2023-11": (60.65, 144.04), "2023-12": (66.77, 145.97),
    "2024-01": (65.83, 126.93), "2024-02": (62.58, 117.64), "2024-03": (62.72, 121.91),
    "2024-04": (59.63, 126.79), "2024-05": (58.92, 136.0),  "2024-06": (57.86, 138.92),
    "2024-07": (57.22, 138.09), "2024-08": (59.58, 133.21), "2024-09": (62.6,  139.9),
    "2024-10": (68.79, 141.83), "2024-11": (73.68, 143.41), "2024-12": (72.92, 141.17),
    "2025-01": (74.17, 134.62), "2025-02": (80.76, 131.57), "2025-03": (89.12, 133.49),
    "2025-04": (83.67, 134.68), "2025-05": (73.3,  133.1),  "2025-06": (68.15, 134.4),
    "2025-07": (63.63, 136.89), "2025-08": (63.87, 140.5),  "2025-09": (64.77, 138.77),
    "2025-10": (65.35, 137.86), "2025-11": (67.54, 140.47), "2025-12": (69.62, 142.01),
    "2026-01": (67.84, 130.98), "2026-02": (67.86, 126.33), "2026-03": (70.9,  129.38),
    "2026-04": (69.74, 128.01),
    # ── Add new months here as needed ────────────────────────────────────────
    # "2026-05": (corn_avg, soy_pna_avg),  # BRL/sc60kg monthly avg from CEPEA/ESALQ
}


# ══════════════════════════════════════════════════════════════════════════════
# DATABASE
# ══════════════════════════════════════════════════════════════════════════════
def init_db(conn):
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS _fx_raw (
        dt   TEXT PRIMARY KEY,
        fx   REAL
    );
    CREATE TABLE IF NOT EXISTS _cepea_grain_raw (
        dt          TEXT PRIMARY KEY,
        corn_brl_sc REAL,
        soy_brl_sc  REAL
    );
    CREATE TABLE IF NOT EXISTS _secex_raw (
        year         INTEGER,
        month        INTEGER,
        rev_000usd   REAL,
        vol_tons     REAL,
        price_usd_kg REAL,
        PRIMARY KEY (year, month)
    );
    CREATE TABLE IF NOT EXISTS _weekly_raw (
        start_date   TEXT PRIMARY KEY,
        end_date     TEXT,
        price_usd_kg REAL,
        vol_tons     REAL,   -- MTD cumulative tons from SECEX weekly bulletin
        biz_days     INTEGER -- business days (Mon-Fri excl. BR holidays) in the period
    );
    CREATE TABLE IF NOT EXISTS monthly (
        period       TEXT PRIMARY KEY,
        year         INTEGER,
        month        INTEGER,
        secex_usd_kg REAL,
        fx           REAL,
        secex_brl_kg REAL,
        cepea_r_kg   REAL,   -- grain basket BRL/kg (2-mo lag)
        spread       REAL,   -- (secex_brl_kg - grain_brl_kg) / secex_brl_kg
        updated_at   TEXT
    );
    CREATE TABLE IF NOT EXISTS weekly (
        start_date     TEXT PRIMARY KEY,
        end_date       TEXT,
        secex_usd_kg   REAL,
        fx             REAL,
        secex_brl_kg   REAL,
        cepea_r_kg     REAL,    -- grain basket BRL/kg (2-mo lag)
        spread         REAL,    -- (secex_brl_kg - grain_brl_kg) / secex_brl_kg
        vol_tons       REAL,    -- incremental weekly tons (de-accumulated from MTD)
        biz_days       INTEGER, -- business days (Mon-Fri excl. BR holidays) in the period
        vol_tons_daily REAL,    -- daily average = vol_tons / biz_days
        updated_at     TEXT
    );
    """)
    conn.commit()
    # ── Migrate existing DBs ───────────────────────────────────────────────────
    migrations = [
        ("_weekly_raw", "vol_tons",     "REAL"),
        ("_weekly_raw", "biz_days",     "INTEGER"),
        ("weekly",      "vol_tons",     "REAL"),
        ("weekly",      "biz_days",     "INTEGER"),
        ("weekly",      "vol_tons_daily", "REAL"),
    ]
    for tbl, col, typ in migrations:
        try:
            conn.execute(f"ALTER TABLE {tbl} ADD COLUMN {col} {typ}")
            conn.commit()
            print(f"  [DB] Migrated: added {tbl}.{col}")
        except Exception:
            pass


# ══════════════════════════════════════════════════════════════════════════════
# HTTP HELPER
# ══════════════════════════════════════════════════════════════════════════════
def get(url, **kwargs):
    hdrs = {"Accept": "application/json", "User-Agent": "Mozilla/5.0"}
    ssl_verify = not any(h in url for h in _NO_VERIFY_HOSTS)
    for attempt in range(RETRY):
        try:
            r = requests.get(url, headers=hdrs, timeout=TIMEOUT,
                             verify=ssl_verify, **kwargs)
            r.raise_for_status()
            return r
        except Exception as e:
            if attempt == RETRY - 1:
                print(f"  ✗ {url[:70]}…: {e}")
                return None
            time.sleep(2 ** attempt)
    return None


# ══════════════════════════════════════════════════════════════════════════════
# BCB PTAX FX
# ══════════════════════════════════════════════════════════════════════════════
def fetch_fx(conn):
    """Download BCB PTAX BRL/USD daily rates into _fx_raw."""
    start = datetime(ANO_INI, 1, 1).strftime("%m-%d-%Y")
    end   = datetime.now().strftime("%m-%d-%Y")
    rows  = []

    # Method A: BCB OLINDA
    url = (
        "https://olinda.bcb.gov.br/olinda/servico/PTAX/versao/v1/odata/"
        f"CotacaoDolarPeriodo(dataInicial=@dataInicial,dataFinalCotacao=@dataFinalCotacao)"
        f"?@dataInicial='{start}'&@dataFinalCotacao='{end}'"
        "&$format=json&$select=cotacaoCompra,dataHoraCotacao"
    )
    r = get(url)
    if r:
        for item in r.json().get("value", []):
            try:
                rows.append((item["dataHoraCotacao"][:10], float(item["cotacaoCompra"])))
            except Exception:
                pass
        print(f"  [FX] BCB OLINDA: {len(rows)} rows")

    # Method B: BCB SGS fallback
    if not rows:
        url2 = (
            f"https://api.bcb.gov.br/dados/serie/bcdata.sgs.1/dados"
            f"?formato=json&dataInicial=01/01/{ANO_INI}&dataFinal="
            + datetime.now().strftime("%d/%m/%Y")
        )
        r2 = get(url2)
        if r2:
            for item in r2.json():
                try:
                    dt = datetime.strptime(item["data"], "%d/%m/%Y").strftime("%Y-%m-%d")
                    rows.append((dt, float(item["valor"])))
                except Exception:
                    pass
            print(f"  [FX] BCB SGS: {len(rows)} rows")

    if not rows:
        print("  [FX] All methods failed — FX not updated.")
        return 0

    conn.executemany("INSERT OR REPLACE INTO _fx_raw(dt,fx) VALUES(?,?)", rows)
    conn.commit()
    print(f"  [FX] {len(rows)} rows stored.")
    return len(rows)


# ══════════════════════════════════════════════════════════════════════════════
# SECEX MONTHLY
# ══════════════════════════════════════════════════════════════════════════════
def fetch_secex(conn, years=None):
    """Download MDIC annual CSVs and upsert into _secex_raw (NCM 0207 poultry).
    Uses only stdlib csv — no pandas required."""
    import csv
    from collections import defaultdict

    if years is None:
        years = range(ANO_INI, datetime.now().year + 1)

    BASE = "https://balanca.economia.gov.br/balanca/bd/comexstat-bd/ncm/EXP_{year}.csv"
    total = 0
    for yr in years:
        r = get(BASE.format(year=yr))
        if not r:
            continue
        try:
            # MDIC CSVs are semicolon-delimited, latin-1 encoded
            text = r.content.decode("latin-1", errors="replace")
            reader = csv.DictReader(text.splitlines(), delimiter=";")

            # Aggregate vol_kg and rev_usd by month for NCM 0207*
            by_month: dict = defaultdict(lambda: [0.0, 0.0])  # month → [vol_kg, rev_usd]
            for row in reader:
                ncm = row.get("CO_NCM", "").strip().zfill(8)
                if ncm[:4] != NCM_CODE:
                    continue
                try:
                    mo      = int(row["CO_MES"])
                    vol_kg  = float(row["KG_LIQUIDO"])
                    rev_usd = float(row["VL_FOB"])
                except (KeyError, ValueError):
                    continue
                by_month[mo][0] += vol_kg
                by_month[mo][1] += rev_usd

            if not by_month:
                print(f"  [SECEX] {yr}: no poultry rows (NCM {NCM_CODE})")
                continue

            rows = []
            for mo, (vol_kg, rev_usd) in sorted(by_month.items()):
                vol = vol_kg  / 1000.0   # kg → tons
                rev = rev_usd / 1000.0   # USD → 000 USD
                p   = (rev_usd / vol_kg) if vol_kg > 0 else None  # USD/kg
                rows.append((yr, mo, rev, vol, p))

            conn.executemany(
                "INSERT OR REPLACE INTO _secex_raw(year,month,rev_000usd,vol_tons,price_usd_kg)"
                " VALUES(?,?,?,?,?)", rows
            )
            conn.commit()
            total += len(rows)
            print(f"  [SECEX] {yr}: {len(rows)} months")
        except Exception as ex:
            print(f"  [SECEX] {yr}: {ex}")
    return total


# ══════════════════════════════════════════════════════════════════════════════
# CEPEA GRAIN (local XLSX — seeding only)
# ══════════════════════════════════════════════════════════════════════════════
def load_cepea_grain(conn, path):
    """
    Load CEPEA corn + soy prices from local xlsx file.
    Accepts either:
      • BZ_Chicken_Support_CLEAN.xlsx (sheet: CEPEA_Daily, cols: Date|Corn|SoyPNA|SoySOR)
      • Any xlsx with columns: date, corn_brl_sc, soy_brl_sc (flexible detection)
    """
    try:
        import openpyxl
    except ImportError:
        sys.exit("Missing: pip install openpyxl")

    p = Path(path)
    if not p.exists():
        print(f"  [CEPEA] Not found: {path}")
        return 0

    rows = []
    try:
        wb = openpyxl.load_workbook(p, data_only=True)

        # Try CEPEA_Daily sheet first (BZ_Chicken_Support_CLEAN.xlsx)
        ws = None
        for candidate in ("CEPEA_Daily", wb.active.title, wb.sheetnames[0]):
            if candidate in wb.sheetnames:
                ws = wb[candidate]
                break

        if ws is None:
            print(f"  [CEPEA] No usable sheet found in {p.name}")
            return 0

        # Detect header row
        # Expected: Date | Corn BRL/sc60kg | Soy PNA BRL/sc60kg | ...
        date_col = corn_col = soy_col = None
        header_row = None
        for row in ws.iter_rows(max_row=5):
            for cell in row:
                v = str(cell.value or "").lower()
                if "date" in v or "data" in v:
                    date_col = cell.column
                    header_row = cell.row
                elif "corn" in v or "milho" in v:
                    corn_col = cell.column
                elif "soy" in v and "pna" in v:
                    soy_col = cell.column
                elif ("soy" in v or "soja" in v) and soy_col is None:
                    soy_col = cell.column
            if date_col is not None:
                break

        if date_col is None:
            # Fallback: assume col 1=date, 2=corn, 3=soy
            date_col, corn_col, soy_col = 1, 2, 3
            header_row = 1

        data_start = (header_row or 1) + 1
        for row in ws.iter_rows(min_row=data_start, values_only=False):
            try:
                dt_cell   = row[date_col - 1].value if date_col else None
                corn_cell = row[corn_col - 1].value if corn_col else None
                soy_cell  = row[soy_col  - 1].value if soy_col  else None

                # Parse date
                if isinstance(dt_cell, (datetime, date)):
                    dt_str = str(dt_cell.date() if isinstance(dt_cell, datetime) else dt_cell)
                elif isinstance(dt_cell, str) and len(dt_cell) >= 8:
                    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
                        try:
                            dt_str = datetime.strptime(dt_cell.strip(), fmt).strftime("%Y-%m-%d")
                            break
                        except Exception:
                            dt_str = None
                    else:
                        continue
                else:
                    continue

                def _float(v):
                    if v is None: return None
                    try: return float(str(v).replace(",", "."))
                    except Exception: return None

                corn = _float(corn_cell)
                soy  = _float(soy_cell)

                if corn is None and soy is None:
                    continue
                # Sanity: corn ~15–120 BRL/sc; soy ~40–220 BRL/sc
                if corn is not None and not (10 < corn < 300):
                    corn = None
                if soy is not None and not (20 < soy < 500):
                    soy = None

                rows.append((dt_str, corn, soy))
            except Exception:
                continue

    except Exception as ex:
        print(f"  [CEPEA] Error reading {p.name}: {ex}")
        return 0

    if not rows:
        print(f"  [CEPEA] No valid rows parsed from {p.name}")
        return 0

    conn.executemany(
        "INSERT OR REPLACE INTO _cepea_grain_raw(dt,corn_brl_sc,soy_brl_sc) VALUES(?,?,?)",
        rows
    )
    conn.commit()
    print(f"  [CEPEA] {len(rows)} rows loaded from {p.name}")
    return len(rows)


# ══════════════════════════════════════════════════════════════════════════════
# WEEKLY RAW SEED
# ══════════════════════════════════════════════════════════════════════════════
def seed_weekly_raw(conn):
    """Seed _weekly_raw from WEEKLY_SEED with biz_days and smart conflict resolution."""
    enriched = []
    for s, e, price, vol_mtd in WEEKLY_SEED:
        bd = _biz_days_between(date.fromisoformat(s), date.fromisoformat(e))
        enriched.append((s, e, price, vol_mtd, bd))

    conn.executemany(
        """
        INSERT INTO _weekly_raw(start_date, end_date, price_usd_kg, vol_tons, biz_days)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(start_date) DO UPDATE SET
            end_date     = excluded.end_date,
            price_usd_kg = excluded.price_usd_kg,
            vol_tons     = excluded.vol_tons,
            biz_days     = excluded.biz_days
        """,
        enriched
    )
    conn.commit()
    print(f"  [WEEKLY] {len(WEEKLY_SEED)} rows seeded/validated in _weekly_raw.")


# ══════════════════════════════════════════════════════════════════════════════
# CEPEA GRAIN SEED  (from embedded CEPEA_MONTHLY_SEED)
# ══════════════════════════════════════════════════════════════════════════════
def seed_cepea_grain(conn):
    """
    Insert monthly CEPEA averages from CEPEA_MONTHLY_SEED into _cepea_grain_raw.
    Uses INSERT OR IGNORE so it never overwrites real daily data already in the DB.
    Each entry is stored as 'YYYY-MM-01' (synthetic first-of-month representative).
    This ensures _grain_cost_brl_kg() always has something to average for
    historical months even when no daily CEPEA rows are present.
    """
    rows = [
        (f"{ym}-01", corn, soy)
        for ym, (corn, soy) in CEPEA_MONTHLY_SEED.items()
    ]
    conn.executemany(
        "INSERT OR IGNORE INTO _cepea_grain_raw(dt, corn_brl_sc, soy_brl_sc) VALUES(?,?,?)",
        rows
    )
    conn.commit()
    print(f"  [CEPEA_SEED] {len(rows)} monthly averages ensured in _cepea_grain_raw.")


# ══════════════════════════════════════════════════════════════════════════════
# FETCH CEPEA DAILY  (scrape cepea.org.br)
# ══════════════════════════════════════════════════════════════════════════════
def fetch_cepea_daily(conn):
    """
    Scrape the last ~15 trading days of CEPEA corn and soy prices from
    cepea.org.br and upsert into _cepea_grain_raw (daily granularity).

    Endpoints:
      • Corn  – https://cepea.org.br/br/indicador/milho.aspx
                Table 0: INDICADOR DO MILHO ESALQ/BM&FBOVESPA (R$/sc60kg)
      • Soy   – https://cepea.org.br/br/indicador/soja.aspx
                Table 0: INDICADOR DA SOJA CEPEA/ESALQ – PARANAGUÁ (R$/sc60kg)

    Uses INSERT OR REPLACE so newer scrapes overwrite stale entries for the
    same date (e.g. provisional → official revision).
    """
    import re as _re, urllib.request as _ur

    CORN_URL = "https://cepea.org.br/br/indicador/milho.aspx"
    SOY_URL  = "https://cepea.org.br/br/indicador/soja.aspx"
    _HDRS = {
        "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0 Safari/537.36",
        "Accept":          "text/html,*/*",
        "Accept-Language": "pt-BR,pt;q=0.9",
        "Referer":         "https://cepea.org.br/",
    }

    def _fetch_html(url):
        req = _ur.Request(url, headers=_HDRS)
        with _ur.urlopen(req, timeout=20) as r:
            return r.read().decode("utf-8", errors="replace")

    def _parse_table(html, table_idx=0):
        """Return list of (ISO-date, price_brl_sc) from the nth HTML table."""
        tables = _re.findall(r"<table[^>]*>(.*?)</table>", html, _re.DOTALL)
        if table_idx >= len(tables):
            return []
        rows = _re.findall(r"<tr[^>]*>(.*?)</tr>", tables[table_idx], _re.DOTALL)
        out = []
        for row in rows:
            cells = _re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", row, _re.DOTALL)
            clean = [_re.sub(r"<[^>]+>", "", c).strip() for c in cells]
            if len(clean) < 2:
                continue
            m = _re.match(r"(\d{2})/(\d{2})/(\d{4})", clean[0])
            if not m:
                continue
            dt_iso = f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
            try:
                price = float(clean[1].replace(".", "").replace(",", "."))
            except ValueError:
                continue
            out.append((dt_iso, price))
        return out

    corn_data: dict = {}
    soy_data:  dict = {}

    for label, url, store, idx in [
        ("Corn ESALQ/BM&F",  CORN_URL, corn_data, 0),
        ("Soy PNA Paranaguá", SOY_URL, soy_data,  0),
    ]:
        try:
            rows = _parse_table(_fetch_html(url), idx)
            for dt, price in rows:
                store[dt] = price
            print(f"  [CEPEA] {label}: {len(rows)} days scraped")
        except Exception as exc:
            print(f"  [CEPEA] {label}: fetch error — {exc}")

    all_dates = set(corn_data) | set(soy_data)
    if not all_dates:
        print("  [CEPEA] No data retrieved — check network / cepea.org.br availability.")
        return 0

    upserted = 0
    for dt in sorted(all_dates):
        corn = corn_data.get(dt)
        soy  = soy_data.get(dt)
        # Only insert rows where at least one price is available
        if corn is None and soy is None:
            continue
        conn.execute(
            "INSERT OR REPLACE INTO _cepea_grain_raw(dt, corn_brl_sc, soy_brl_sc) VALUES(?,?,?)",
            (dt, corn, soy),
        )
        upserted += 1

    conn.commit()
    print(f"  [CEPEA] {upserted} daily rows upserted into _cepea_grain_raw.")
    return upserted


# ══════════════════════════════════════════════════════════════════════════════
# FETCH WEEKLY SECEX BULLETIN
# ══════════════════════════════════════════════════════════════════════════════
def fetch_weekly_bulletin(conn):
    """
    Fetch MDIC weekly bulletin from the fixed URL:
      https://balanca.economia.gov.br/balanca/semanal/Setores_Produtos.xlsx

    This file is updated every week and always holds the current MTD cumulative
    data for exports by sector/product (CUCI).

    File structure (sheet EXP):
      Row 7, col 2 : period header e.g. "Abr/2026"
      Col 1        : product description
      Col 6        : Toneladas MTD (cumulative from 1st of month)
      Col 10       : Preço US$/Tonelada → divide by 1000 for USD/kg

    vol_tons stored as MTD cumulative; materialise() de-accumulates monthly.
    """
    import io, re as _re, zipfile as _zf
    from datetime import date as _date, timedelta as _td

    XLSX_URL   = "https://balanca.economia.gov.br/balanca/semanal/Setores_Produtos.xlsx"
    PT_MON_REV = {"jan":1,"fev":2,"mar":3,"abr":4,"mai":5,"jun":6,
                  "jul":7,"ago":8,"set":9,"out":10,"nov":11,"dez":12}
    POULTRY_KW = ("aves", "frango", "galinha", "peru", "pato", "0207")
    today = _date.today()

    # ── Download ──────────────────────────────────────────────────────────────
    print(f"  [BULLETIN] Downloading Setores_Produtos.xlsx …")
    r = get(XLSX_URL)
    if r is None:
        return 0

    # ── Load (MDIC files always have drawing refs — strip them) ───────────────
    try:
        import openpyxl
        raw = r.content
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        except Exception:
            buf = io.BytesIO()
            with _zf.ZipFile(io.BytesIO(raw), 'r') as zin:
                with _zf.ZipFile(buf, 'w', _zf.ZIP_DEFLATED) as zout:
                    for item in zin.infolist():
                        if 'drawing' in item.filename.lower():
                            continue
                        data = zin.read(item.filename)
                        if item.filename.endswith('.rels'):
                            data = _re.sub(
                                rb'<Relationship[^>]+/drawing[^>]+/?>', b'', data)
                        zout.writestr(item, data)
            buf.seek(0)
            wb = openpyxl.load_workbook(buf, data_only=True)
    except Exception as exc:
        print(f"  [BULLETIN] Excel load error: {exc}")
        return 0

    ws = wb['EXP'] if 'EXP' in wb.sheetnames else wb.active

    # ── Parse period from header rows (e.g. "Abr/2026") ──────────────────────
    bull_year = bull_month = None
    for ri in range(1, 12):
        for c in range(1, min(ws.max_column + 1, 16)):
            v = str(ws.cell(ri, c).value or "")
            m = _re.search(
                r'(jan|fev|mar|abr|mai|jun|jul|ago|set|out|nov|dez)[a-z]*/(\d{4})',
                v, _re.IGNORECASE)
            if m:
                bull_month = PT_MON_REV[m.group(1).lower()[:3]]
                bull_year  = int(m.group(2))
                break
        if bull_year:
            break

    if bull_year is None:
        print("  [BULLETIN] Could not parse period from file header.")
        return 0

    period_str = f"{bull_year}-{bull_month:02d}"
    print(f"  [BULLETIN] Period: {period_str}")

    # ── Find poultry row ──────────────────────────────────────────────────────
    # Column layout (confirmed from live file):
    #   col 1  = Descrição
    #   col 2  = US$ Mil (MTD current year)
    #   col 6  = Toneladas (MTD current year)  ← vol_tons
    #   col 10 = Preço US$/Tonelada            ← price (÷1000 → USD/kg)
    bull_price_usd_kg = bull_vol_tons = None
    for ri in range(1, ws.max_row + 1):
        desc = str(ws.cell(ri, 1).value or "").lower()
        if any(kw in desc for kw in POULTRY_KW):
            try:
                vol   = float(ws.cell(ri, 6).value  or 0)
                price = float(ws.cell(ri, 10).value or 0)
                if vol > 0 and price > 0:
                    bull_vol_tons     = vol
                    bull_price_usd_kg = price / 1000.0
                    print(f"  [BULLETIN] '{ws.cell(ri,1).value}'[:55]: "
                          f"vol_mtd={vol:.0f} t, price={bull_price_usd_kg:.4f} USD/kg")
            except Exception:
                pass
            break

    if bull_vol_tons is None:
        print("  [BULLETIN] Poultry row not found or zero data.")
        return 0

    # ── Guard: skip if vol already covered ───────────────────────────────────
    last_row = conn.execute(
        "SELECT start_date, end_date, vol_tons FROM _weekly_raw "
        "WHERE strftime('%Y-%m', start_date) = ? ORDER BY start_date DESC LIMIT 1",
        (period_str,)
    ).fetchone()

    if last_row and last_row[2] is not None and bull_vol_tons <= last_row[2] + 1.0:
        print(f"  [BULLETIN] Already up to date for {period_str} "
              f"(bulletin={bull_vol_tons:.0f} t ≤ stored={last_row[2]:.0f} t).")
        return 0

    # ── Date range for the new entry ─────────────────────────────────────────
    # start = day after last stored end_date for this month (or 1st of month)
    # end   = today
    last_end = conn.execute(
        "SELECT MAX(end_date) FROM _weekly_raw "
        "WHERE strftime('%Y-%m', start_date) = ?",
        (period_str,)
    ).fetchone()[0]

    if last_end:
        week_start = (_date.fromisoformat(last_end) + _td(days=1)).isoformat()
    else:
        week_start = f"{bull_year}-{bull_month:02d}-01"
    week_end = str(today)

    biz = _biz_days_between(_date.fromisoformat(week_start), _date.fromisoformat(week_end))

    # Safety: never let the bulletin overwrite a SEED row (start_date already in DB
    # from a prior hardcoded entry).  Seeds always win; the bulletin should only
    # INSERT new rows whose start_date is strictly after any existing end_date.
    existing = conn.execute(
        "SELECT vol_tons FROM _weekly_raw WHERE start_date=?", (week_start,)
    ).fetchone()
    if existing is not None:
        # The date slot is already taken by a seed row — push week_start forward
        # to the day after that row's end_date.
        taken_end = conn.execute(
            "SELECT end_date FROM _weekly_raw WHERE start_date=?", (week_start,)
        ).fetchone()[0]
        week_start = (_date.fromisoformat(taken_end) + _td(days=1)).isoformat()
        biz = _biz_days_between(_date.fromisoformat(week_start),
                                _date.fromisoformat(week_end))
        print(f"  [BULLETIN] Seed conflict — adjusted week_start to {week_start}")

    conn.execute(
        "INSERT INTO _weekly_raw(start_date, end_date, price_usd_kg, vol_tons, biz_days) "
        "VALUES (?,?,?,?,?) "
        "ON CONFLICT(start_date) DO UPDATE SET "
        "  end_date=excluded.end_date, "
        "  price_usd_kg=COALESCE(excluded.price_usd_kg, _weekly_raw.price_usd_kg), "
        "  vol_tons=excluded.vol_tons, biz_days=excluded.biz_days",
        (week_start, week_end, bull_price_usd_kg, bull_vol_tons, biz),
    )
    conn.commit()
    print(f"  [BULLETIN] Upserted {week_start}→{week_end}: "
          f"price={bull_price_usd_kg:.4f} USD/kg, vol={bull_vol_tons:.0f} t MTD, "
          f"biz_days={biz}")
    return 1


# ══════════════════════════════════════════════════════════════════════════════
# FILL SECEX FROM WEEKLY
# ══════════════════════════════════════════════════════════════════════════════
def fill_secex_from_weekly(conn):
    """
    For the current (incomplete) month, estimate monthly SECEX from weekly data
    if SECEX monthly hasn't published yet.
    Only inserts if that month is not already in _secex_raw.
    """
    from datetime import date as _date
    today = _date.today()
    yr, mo = today.year, today.month
    exists = conn.execute(
        "SELECT 1 FROM _secex_raw WHERE year=? AND month=?", (yr, mo)
    ).fetchone()
    if exists:
        return

    rows = conn.execute(
        """
        SELECT price_usd_kg, vol_tons
        FROM   _weekly_raw
        WHERE  strftime('%Y', start_date) = ?
          AND  strftime('%m', start_date) = ?
          AND  price_usd_kg IS NOT NULL
          AND  vol_tons      IS NOT NULL
        """,
        (str(yr), f"{mo:02d}")
    ).fetchall()

    if not rows:
        return

    total_vol = sum(r[1] for r in rows)
    if total_vol == 0:
        return
    avg_price  = sum(r[0] * r[1] for r in rows) / total_vol
    rev_000usd = avg_price * total_vol

    conn.execute(
        """
        INSERT OR IGNORE INTO _secex_raw(year, month, rev_000usd, vol_tons, price_usd_kg)
        VALUES (?,?,?,?,?)
        """,
        (yr, mo, rev_000usd, total_vol, avg_price),
    )
    conn.commit()
    print(f"  [WEEKLY→SECEX] Estimated {yr}-{mo:02d} from weekly: "
          f"price={avg_price:.4f} USD/kg, vol={total_vol:.0f} t")


# ══════════════════════════════════════════════════════════════════════════════
# GRAIN COST HELPER
# ══════════════════════════════════════════════════════════════════════════════
def _grain_cost_brl_kg(conn, year, month):
    """
    Return grain basket BRL/kg for (year, month) using a 2-month lag.
    Averages _cepea_grain_raw over the lagged month. Falls back to nearest prior.
    """
    lag_year  = year  if month > GRAIN_LAG else year - 1
    lag_month = month - GRAIN_LAG if month > GRAIN_LAG else month - GRAIN_LAG + 12
    prefix    = f"{lag_year}-{lag_month:02d}"

    row = conn.execute(
        "SELECT AVG(corn_brl_sc), AVG(soy_brl_sc) FROM _cepea_grain_raw WHERE dt LIKE ?",
        (prefix + "%",)
    ).fetchone()

    if row and row[0] is not None and row[1] is not None:
        corn_avg, soy_avg = row
    else:
        fb = conn.execute(
            """
            SELECT corn_brl_sc, soy_brl_sc FROM _cepea_grain_raw
            WHERE  dt < ? AND corn_brl_sc IS NOT NULL AND soy_brl_sc IS NOT NULL
            ORDER BY dt DESC LIMIT 1
            """,
            (prefix + "-99",)
        ).fetchone()
        if fb:
            corn_avg, soy_avg = fb
        else:
            return None

    basket_brl_sc = CORN_WEIGHT * corn_avg + SOY_WEIGHT * soy_avg
    return basket_brl_sc / 60.0


# ══════════════════════════════════════════════════════════════════════════════
# MATERIALISE
# ══════════════════════════════════════════════════════════════════════════════
def materialise(conn):
    """Rebuild monthly and weekly output tables from raw data."""
    now = datetime.utcnow().isoformat()

    # ── MONTHLY ──────────────────────────────────────────────────────────────
    secex_rows = conn.execute(
        "SELECT year, month, price_usd_kg FROM _secex_raw ORDER BY year, month"
    ).fetchall()

    monthly_out = []
    for yr, mo, price_usd_kg in secex_rows:
        if price_usd_kg is None:
            continue
        period = f"{yr}-{mo:02d}"

        fx_row = conn.execute(
            "SELECT AVG(fx) FROM _fx_raw WHERE dt LIKE ?",
            (f"{yr}-{mo:02d}%",)
        ).fetchone()
        fx = fx_row[0] if fx_row and fx_row[0] else None
        if fx is None:
            continue

        secex_brl_kg = price_usd_kg * fx
        cepea_r_kg   = _grain_cost_brl_kg(conn, yr, mo)
        spread = None
        if cepea_r_kg is not None and secex_brl_kg:
            spread = (secex_brl_kg - cepea_r_kg) / secex_brl_kg

        monthly_out.append((period, yr, mo, price_usd_kg, fx, secex_brl_kg,
                            cepea_r_kg, spread, now))

    conn.executemany(
        """
        INSERT OR REPLACE INTO monthly
          (period, year, month, secex_usd_kg, fx, secex_brl_kg,
           cepea_r_kg, spread, updated_at)
        VALUES (?,?,?,?,?,?,?,?,?)
        """,
        monthly_out,
    )
    conn.commit()
    print(f"  [MAT] monthly: {len(monthly_out)} rows")

    # ── WEEKLY ───────────────────────────────────────────────────────────────
    weekly_rows = conn.execute(
        "SELECT start_date, end_date, price_usd_kg, vol_tons, biz_days"
        " FROM _weekly_raw ORDER BY start_date"
    ).fetchall()

    # De-accumulate MTD volumes within each calendar month
    prev_month = None
    prev_vol   = 0.0
    deacc = []
    for start_date, end_date, price_usd_kg, vol_mtd, biz_d in weekly_rows:
        yr_mo = start_date[:7]
        if yr_mo != prev_month:
            prev_month = yr_mo
            prev_vol   = 0.0
        inc_vol = None
        if vol_mtd is not None:
            inc_vol  = max(0.0, vol_mtd - prev_vol)
            prev_vol = vol_mtd
        # Compute biz_days from dates if not stored
        if biz_d is None:
            biz_d = _biz_days_between(
                date.fromisoformat(start_date),
                date.fromisoformat(end_date or start_date)
            )
        deacc.append((start_date, end_date, price_usd_kg, inc_vol, biz_d))

    weekly_out = []
    for start_date, end_date, price_usd_kg, inc_vol, biz_d in deacc:
        if price_usd_kg is None:
            continue
        yr = int(start_date[:4])
        mo = int(start_date[5:7])
        ed = end_date or start_date

        fx_row = conn.execute(
            "SELECT AVG(fx) FROM _fx_raw WHERE dt >= ? AND dt <= ?",
            (start_date, ed)
        ).fetchone()
        fx = fx_row[0] if fx_row and fx_row[0] else None
        if fx is None:
            fx_row2 = conn.execute(
                "SELECT AVG(fx) FROM _fx_raw WHERE dt LIKE ?",
                (f"{yr}-{mo:02d}%",)
            ).fetchone()
            fx = fx_row2[0] if fx_row2 and fx_row2[0] else None
        if fx is None:
            continue

        secex_brl_kg = price_usd_kg * fx
        cepea_r_kg   = _grain_cost_brl_kg(conn, yr, mo)
        spread = None
        if cepea_r_kg is not None and secex_brl_kg:
            spread = (secex_brl_kg - cepea_r_kg) / secex_brl_kg

        vol_daily = (round(inc_vol / biz_d, 3)
                     if inc_vol is not None and biz_d and biz_d > 0
                     else None)

        weekly_out.append((start_date, end_date, price_usd_kg, fx, secex_brl_kg,
                           cepea_r_kg, spread, inc_vol, biz_d, vol_daily, now))

    conn.executemany(
        """
        INSERT OR REPLACE INTO weekly
          (start_date, end_date, secex_usd_kg, fx, secex_brl_kg,
           cepea_r_kg, spread, vol_tons, biz_days, vol_tons_daily, updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
        """,
        weekly_out,
    )
    conn.commit()
    print(f"  [MAT] weekly: {len(weekly_out)} rows")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    import argparse
    parser = argparse.ArgumentParser(description="Refresh chicken_bz.db")
    parser.add_argument("--init",  action="store_true",
                        help="Full reseed (all years) — use for first-time setup")
    parser.add_argument("--cepea", metavar="PATH",
                        help="Optional: seed CEPEA grain from xlsx file")
    args = parser.parse_args()

    print(f"[DB] Opening {DB_PATH}")
    conn = sqlite3.connect(DB_PATH)
    init_db(conn)

    # ── 1. Weekly historical seed ─────────────────────────────────────────────
    seed_weekly_raw(conn)

    # ── 2. CEPEA grain ────────────────────────────────────────────────────────
    seed_cepea_grain(conn)
    fetch_cepea_daily(conn)
    if args.cepea:
        load_cepea_grain(conn, args.cepea)

    # ── 3. SECEX monthly ──────────────────────────────────────────────────────
    if args.init:
        print("[SECEX] Full reseed — all years …")
        fetch_secex(conn)
    else:
        cur_year = datetime.now().year
        fetch_secex(conn, years=[cur_year - 1, cur_year])

    # ── 4. BCB PTAX FX ───────────────────────────────────────────────────────
    fetch_fx(conn)

    # ── 5. Weekly bulletin + fill ─────────────────────────────────────────────
    fetch_weekly_bulletin(conn)
    fill_secex_from_weekly(conn)

    # ── 6. Materialise ────────────────────────────────────────────────────────
    materialise(conn)

    conn.close()
    db_size = DB_PATH.stat().st_size // 1024
    print(f"\n✓ Done. {DB_PATH.name} = {db_size} KB")


if __name__ == "__main__":
    main()
                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                 