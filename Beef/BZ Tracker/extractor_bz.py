#!/usr/bin/env python3
"""
extractor_bz.py — Brazil Beef Export Spread Tracker
=====================================================
Builds / refreshes  beef_bz.db  with monthly and weekly spread data.

DATA SOURCES
  • SECEX/MDIC monthly exports  → balanca.economia.gov.br annual CSVs
                                   NCM 0201 (fresh) + 0202 (frozen)
  • BCB PTAX BRL/USD daily FX   → BCB OLINDA API (primary) / BCB SGS (fallback)
  • CEPEA Boi Gordo R$/arroba   → local XLS file (seeding only)
                                   stored in DB, not re-fetched on weekly runs

USAGE
  pip install requests openpyxl
  python extractor_bz.py                          # incremental (SECEX + FX only)
  python extractor_bz.py --init                   # full seed (needs CEPEA XLS)
  python extractor_bz.py --init --cepea PATH.xls  # explicit CEPEA file path

OUTPUT
  beef_bz.db  (SQLite)

SCHEMA — table: monthly
  period       TEXT PRIMARY KEY   e.g. "2006-01"
  year         INTEGER
  month        INTEGER
  secex_usd_kg REAL               FOB USD / kg net weight
  fx           REAL               avg BCB PTAX BRL/USD for the month
  secex_brl_kg REAL               secex_usd_kg × fx
  cepea_r_kg   REAL               avg CEPEA R$/arroba ÷ 15 for the month
  spread       REAL               secex_brl_kg / cepea_r_kg
  updated_at   TEXT               ISO timestamp

SCHEMA — table: weekly
  start_date   TEXT PRIMARY KEY   ISO date "YYYY-MM-DD"
  end_date     TEXT               ISO date "YYYY-MM-DD"
  secex_usd_kg REAL               from SECEX weekly cumulative data
  fx           REAL               avg BCB PTAX BRL/USD for the period
  secex_brl_kg REAL               secex_usd_kg × fx
  cepea_r_kg   REAL               avg CEPEA R$/arroba ÷ 15 for the period
  spread       REAL               secex_brl_kg / cepea_r_kg
  updated_at   TEXT               ISO timestamp
"""

import sqlite3, os, sys, time, warnings
from datetime import datetime, date
from calendar import monthrange
from pathlib import Path

try:
    import requests
    # Suppress InsecureRequestWarning for Brazilian government domains
    # (balanca.economia.gov.br uses ICP-Brasil cert not trusted by default on Linux)
    from urllib3.exceptions import InsecureRequestWarning
    warnings.filterwarnings("ignore", category=InsecureRequestWarning)
except ImportError:
    sys.exit("Missing: pip install requests")

# Domains that require SSL verification disabled (ICP-Brasil / SERPRO chain)
_NO_VERIFY_HOSTS = ("balanca.economia.gov.br", "olinda.bcb.gov.br", "api.bcb.gov.br")

# ── Paths ──────────────────────────────────────────────────────────────────────
DB_PATH    = Path(__file__).parent / "beef_bz.db"
TIMEOUT    = 30
RETRY      = 3
NCM_CODES  = {"0201", "0202"}
ANO_INI    = 2006   # earliest SECEX year to seed

# ── Weekly historical data (Aug 2022 – Mar 2026, 176 weeks) ───────────────────
# Format: (start_date, end_date, price_usd_kg, vol_tons_mtd, rev_000usd_mtd)
#
# price_usd_kg  = exact incremental weekly price = rev_week / vol_week
#                 sourced from Spread_Exportacao_Bovina_Brasil.xlsx (SECEX Weekly tab)
#                 NO rounding — 6 decimal places from actual SECEX bulletin values
# vol_tons_mtd  = MTD cumulative tons within the calendar month
# rev_000usd_mtd= MTD cumulative revenue (000 USD) within the calendar month
#
# materialise() de-accumulates both vol and rev to compute the true weekly
# incremental price:  secex_usd_kg = (rev_mtd - prev_rev_mtd) / (vol_mtd - prev_vol_mtd)
WEEKLY_SEED = [
    # ── 2022 ────────────────────────────────────────────────────────────────────
    ("2022-08-01","2022-08-05",6.246868,39395.1,246096.0),
    ("2022-08-08","2022-08-12",6.312857,88806.1,558020.6),
    ("2022-08-15","2022-08-19",6.191538,128548.7,804088.4),
    ("2022-08-22","2022-08-31",5.921756,203230.5,1246335.8),
    ("2022-09-01","2022-09-09",6.062413,63676.0,386030.2),
    ("2022-09-12","2022-09-16",5.895305,114065.2,683089.9),
    ("2022-09-19","2022-09-23",5.957723,155047.6,927251.7),
    ("2022-09-26","2022-09-30",6.068005,203023.8,1218371.5),
    ("2022-10-03","2022-10-07",5.964508,55618.4,331736.4),
    ("2022-10-10","2022-10-15",5.880332,91159.1,540727.5),
    ("2022-10-17","2022-10-21",5.839897,142692.9,841679.6),
    ("2022-10-24","2022-10-31",5.708353,188557.9,1103493.2),
    ("2022-11-01","2022-11-14",5.350202,68876.2,368501.6),
    ("2022-11-14","2022-11-18",5.155034,98389.0,520641.1),
    ("2022-11-21","2022-11-30",5.101745,148843.6,778047.6),
    ("2022-12-01","2022-12-09",5.032311,47776.6,240426.7),
    ("2022-12-12","2022-12-16",4.921427,81915.5,408438.8),
    ("2022-12-19","2022-12-23",4.89315,116634.2,578322.6),
    ("2022-12-26","2022-12-30",4.926103,152797.9,756468.7),
    # ── 2023 ────────────────────────────────────────────────────────────────────
    ("2023-01-02","2023-01-06",4.880562,40795.9,199106.9),
    ("2023-01-09","2023-01-13",4.854745,77824.6,378871.8),
    ("2023-01-16","2023-01-20",4.832835,107095.6,520333.7),
    ("2023-01-23","2023-01-27",4.811144,160191.1,775783.8),
    ("2023-02-01","2023-02-10",4.822917,47166.5,227480.1),
    ("2023-02-13","2023-02-17",4.891501,91816.6,445886.1),
    ("2023-02-20","2023-02-28",4.852416,126449.7,613940.3),
    ("2023-03-01","2023-03-10",4.869812,67427.1,328357.3),
    ("2023-03-13","2023-03-17",4.845206,89807.1,436793.0),
    ("2023-03-20","2023-03-31",4.631053,107470.6,518593.6),
    ("2023-04-03","2023-04-07",4.546823,22594.5,102733.2),
    ("2023-04-10","2023-04-14",4.719671,43245.6,200199.6),
    ("2023-04-17","2023-04-21",4.826833,72445.1,341140.7),
    ("2023-04-24","2023-04-28",4.935857,110339.9,528184.0),
    ("2023-05-01","2023-05-05",5.062274,42809.4,216712.9),
    ("2023-05-08","2023-05-12",5.06834,75017.4,379954.0),
    ("2023-05-15","2023-05-19",5.116545,113818.7,578482.6),
    ("2023-05-22","2023-05-31",5.133046,168509.6,859213.5),
    ("2023-06-01","2023-06-09",5.144507,70322.5,361774.6),
    ("2023-06-12","2023-06-16",5.060929,114541.5,585563.8),
    ("2023-06-19","2023-06-23",4.989387,154625.3,785557.4),
    ("2023-06-26","2023-06-30",4.947414,192741.6,974134.5),
    ("2023-07-03","2023-07-07",4.862821,37378.3,181764.0),
    ("2023-07-10","2023-07-14",4.801894,76629.1,370242.2),
    ("2023-07-17","2023-07-21",4.612821,117590.0,559187.5),
    ("2023-07-24","2023-07-31",4.699262,160795.4,762221.0),
    ("2023-08-01","2023-08-04",4.531879,41267.3,187018.4),
    ("2023-08-07","2023-08-11",4.482386,83343.8,375621.5),
    ("2023-08-14","2023-08-18",4.525305,124650.1,562545.1),
    ("2023-08-21","2023-08-31",4.506433,185364.9,836152.3),
    ("2023-09-01","2023-09-08",4.488357,74840.6,335911.3),
    ("2023-09-11","2023-09-15",4.529039,119984.0,540367.5),
    ("2023-09-18","2023-09-22",4.594888,158835.7,718886.7),
    ("2023-09-25","2023-09-29",4.583174,195071.7,884962.6),
    ("2023-10-02","2023-10-06",4.591707,38119.9,175035.4),
    ("2023-10-09","2023-10-13",4.62367,91285.7,420856.5),
    ("2023-10-16","2023-10-20",4.54944,133591.5,613324.2),
    ("2023-10-23","2023-10-31",4.610328,186203.9,855884.6),
    ("2023-11-01","2023-11-10",4.598787,73205.0,336654.2),
    ("2023-11-13","2023-11-17",4.555214,119027.5,545385.5),
    ("2023-11-20","2023-11-30",4.612921,187976.8,863443.2),
    ("2023-12-01","2023-12-08",4.590405,64877.5,297814.0),
    ("2023-12-11","2023-12-15",4.538742,97395.3,445403.9),
    ("2023-12-18","2023-12-22",4.526689,166130.4,756546.3),
    ("2023-12-25","2023-12-29",4.523777,208439.4,947942.8),
    # ── 2024 ────────────────────────────────────────────────────────────────────
    ("2024-01-01","2024-01-05",4.517245,49835.2,225117.8),
    ("2024-01-08","2024-01-12",4.542524,86833.4,393183.0),
    ("2024-01-15","2024-01-19",4.465027,123021.6,554764.3),
    ("2024-01-22","2024-01-26",4.505636,168103.2,757885.6),
    ("2024-01-29","2024-01-31",4.702931,181690.3,821784.8),
    ("2024-02-01","2024-02-09",4.577006,50220.1,229857.7),
    ("2024-02-12","2024-02-23",4.519596,143478.2,651346.6),
    ("2024-02-26","2024-02-29",4.474485,179119.6,810823.5),
    ("2024-03-01","2024-03-08",4.501065,50612.2,227808.8),
    ("2024-03-11","2024-03-15",4.512225,84673.8,381502.4),
    ("2024-03-18","2024-03-22",4.557655,139942.0,633395.8),
    ("2024-03-25","2024-03-29",4.545138,166327.6,753322.0),
    ("2024-04-01","2024-04-05",4.482104,54698.2,245163.0),
    ("2024-04-08","2024-04-12",4.547096,104326.6,470828.1),
    ("2024-04-15","2024-04-19",4.558743,155943.2,706134.9),
    ("2024-04-22","2024-04-26",4.534966,203839.7,923343.9),
    ("2024-04-29","2024-04-30",4.616338,208053.3,942795.3),
    ("2024-05-01","2024-05-10",4.491147,75405.9,338659.0),
    ("2024-05-13","2024-05-31",4.512215,211976.0,954892.7),
    ("2024-06-03","2024-06-07",4.460626,57944.2,258467.4),
    ("2024-06-10","2024-06-14",4.445418,97254.7,433219.0),
    ("2024-06-17","2024-06-21",4.499435,146293.5,653865.9),
    ("2024-06-24","2024-06-28",4.456661,192571.4,860110.8),
    ("2024-07-01","2024-07-05",4.436409,54204.0,240471.1),
    ("2024-07-08","2024-07-12",4.394018,109584.0,483811.8),
    ("2024-07-15","2024-07-19",4.413056,162524.5,717441.2),
    ("2024-07-22","2024-07-26",4.423703,215619.3,952316.8),
    ("2024-07-29","2024-07-31",4.332145,237267.1,1046098.2),
    ("2024-08-01","2024-08-09",4.421802,71371.7,315591.5),
    ("2024-08-12","2024-08-16",4.428134,109718.8,485397.6),
    ("2024-08-19","2024-08-23",4.508507,164042.3,730315.5),
    ("2024-08-26","2024-08-30",4.379146,217458.7,964233.7),
    ("2024-09-02","2024-09-06",4.410973,70984.2,313109.4),
    ("2024-09-09","2024-09-13",4.485247,139185.3,619008.2),
    ("2024-09-16","2024-09-20",4.571465,185486.8,830673.9),
    ("2024-09-23","2024-09-27",4.612091,251755.9,1136313.0),
    ("2024-10-01","2024-10-04",4.595323,39866.9,183201.3),
    ("2024-10-07","2024-10-11",4.603673,101660.9,467680.7),
    ("2024-10-14","2024-10-18",4.606428,176399.1,811956.8),
    ("2024-10-21","2024-10-25",4.74176,236196.9,1095503.6),
    ("2024-10-28","2024-10-31",4.814609,270332.3,1259852.2),
    ("2024-11-04","2024-11-08",4.822384,73522.6,354554.2),
    ("2024-11-11","2024-11-15",4.835616,137340.3,663152.1),
    ("2024-11-18","2024-11-22",4.978115,179991.3,875473.7),
    ("2024-11-25","2024-11-29",4.871279,228132.5,1109982.9),
    ("2024-12-02","2024-12-06",4.935534,43033.6,212393.8),
    ("2024-12-09","2024-12-13",4.880585,89335.1,438372.2),
    ("2024-12-16","2024-12-20",5.057792,127290.3,630341.7),
    ("2024-12-23","2024-12-31",4.953413,202569.2,1003229.2),
    # ── 2025 ────────────────────────────────────────────────────────────────────
    ("2025-01-02","2025-01-10",5.057868,66397.7,335830.8),
    ("2025-01-13","2025-01-17",5.014799,112731.7,568186.5),
    ("2025-01-20","2025-01-24",5.029461,143317.4,722016.1),
    ("2025-01-27","2025-01-31",4.994006,180473.7,907574.9),
    ("2025-02-03","2025-02-07",4.959981,47385.5,235031.2),
    ("2025-02-10","2025-02-14",4.937712,99848.6,494078.9),
    ("2025-02-17","2025-02-21",4.904161,153143.1,755443.7),
    ("2025-02-24","2025-02-28",4.904887,190457.8,938468.1),
    ("2025-03-03","2025-03-07",4.892488,60545.0,296215.7),
    ("2025-03-10","2025-03-14",4.86173,117480.6,573021.2),
    ("2025-03-17","2025-03-21",4.911743,163297.9,798064.0),
    ("2025-03-24","2025-03-31",4.954812,215427.2,1056354.9),
    ("2025-04-01","2025-04-04",4.948969,37420.6,185193.4),
    ("2025-04-07","2025-04-11",4.974897,98194.2,487535.8),
    ("2025-04-14","2025-04-17",5.041058,159327.8,795713.8),
    ("2025-04-21","2025-04-25",5.103634,211548.1,1062227.1),
    ("2025-04-28","2025-04-30",5.094358,241583.8,1215239.7),
    ("2025-05-02","2025-05-09",5.103308,67165.3,342765.2),
    ("2025-05-12","2025-05-16",5.130963,123005.5,629279.2),
    ("2025-05-19","2025-05-23",5.32734,173804.1,899900.6),
    ("2025-05-26","2025-05-30",5.291993,218073.7,1134175.0),
    ("2025-06-02","2025-06-06",5.367196,64225.3,344709.8),
    ("2025-06-09","2025-06-13",5.4649,117245.9,634462.1),
    ("2025-06-16","2025-06-20",5.477433,168837.9,917053.8),
    ("2025-06-23","2025-06-30",5.487408,241098.7,1313578.3),
    ("2025-07-01","2025-07-04",5.540986,48715.5,269931.9),
    ("2025-07-07","2025-07-11",5.531584,104193.7,576814.2),
    ("2025-07-14","2025-07-18",5.566097,172709.5,958179.8),
    ("2025-07-21","2025-07-25",5.539654,243904.9,1352577.7),
    ("2025-07-28","2025-07-31",5.591479,276879.0,1536951.7),
    ("2025-08-01","2025-08-08",5.557195,80470.4,447189.7),
    ("2025-08-11","2025-08-15",5.734543,135785.1,764394.2),
    ("2025-08-18","2025-08-22",5.553723,212925.3,1192809.5),
    ("2025-08-25","2025-08-29",5.594571,268562.6,1504076.3),
    ("2025-09-01","2025-09-05",5.555435,78338.9,435206.7),
    ("2025-09-08","2025-09-12",5.699694,137274.5,771121.6),
    ("2025-09-15","2025-09-19",5.64287,209645.3,1179500.6),
    ("2025-09-22","2025-09-26",5.581394,294706.7,1654261.8),
    ("2025-09-29","2025-09-30",5.67875,314689.9,1767741.4),
    ("2025-10-01","2025-10-10",5.551596,111919.9,621334.1),
    ("2025-10-13","2025-10-17",5.449638,201346.8,1108678.3),
    ("2025-10-20","2025-10-24",5.57805,276493.4,1527849.8),
    ("2025-10-27","2025-10-31",5.620744,320559.4,1775533.5),
    ("2025-11-03","2025-11-07",5.51078,100536.4,554034.0),
    ("2025-11-10","2025-11-14",5.557454,163699.5,905060.0),
    ("2025-11-17","2025-11-21",5.408732,238219.7,1308119.8),
    ("2025-11-24","2025-11-28",5.560989,318493.4,1754521.0),
    ("2025-12-01","2025-12-05",5.617311,76721.3,430967.4),
    ("2025-12-08","2025-12-12",5.587754,143577.8,804545.1),
    ("2025-12-15","2025-12-19",5.563918,218356.4,1220607.1),
    ("2025-12-22","2025-12-31",5.646376,304977.1,1709700.1),
    # ── 2026 ────────────────────────────────────────────────────────────────────
    ("2026-01-05","2026-01-09",5.529294,89307.3,493806.3),
    ("2026-01-12","2026-01-16",5.579579,126254.1,699953.9),
    ("2026-01-19","2026-01-23",5.64871,183782.9,1024917.4),
    ("2026-01-26","2026-01-30",5.55938,231821.3,1291981.1),
    ("2026-02-02","2026-02-06",5.619449,68344.0,384055.6),
    ("2026-02-09","2026-02-13",5.570207,136800.0,765369.7),
    ("2026-02-16","2026-02-20",5.658724,192708.7,1081741.6),
    ("2026-02-23","2026-02-27",5.76357,235889.7,1330618.3),
    ("2026-03-02","2026-03-06",5.687826,59986.7,341193.9),
    ("2026-03-09","2026-03-13",5.848168,115678.5,666888.9),
    ("2026-03-16","2026-03-20",5.825231,167061.8,966208.5),
    ("2026-03-23","2026-03-31",5.892906,233951.5,1360383.2),
    # ── 2026-Apr ─────────────────────────────────────────────────────────────────
    ("2026-04-01","2026-04-10",6.078718,97264.669,591244.475),   # biz_days=7
    ("2026-04-11","2026-04-17",6.255470,153353.3,942105.2),      # biz_days=5
    ("2026-04-20","2026-04-24",6.340334,216266.444,1340995.538), # biz_days=4 (Mon–Thu)
]


# ══════════════════════════════════════════════════════════════════════════════
# DATABASE
# ══════════════════════════════════════════════════════════════════════════════
def init_db(conn):
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS _fx_raw (
        dt   TEXT PRIMARY KEY,
        fx   REAL
    );
    CREATE TABLE IF NOT EXISTS _cepea_raw (
        dt        TEXT PRIMARY KEY,
        r_arroba  REAL,
        r_kg      REAL
    );
    CREATE TABLE IF NOT EXISTS _secex_raw (
        year         INTEGER,
        month        INTEGER,
        rev_000usd   REAL,
        vol_tons     REAL,
        price_usd_kg REAL,
        PRIMARY KEY (year, month)
    );
    CREATE TABLE IF NOT EXISTS _weekly_raw (
        start_date   TEXT PRIMARY KEY,
        end_date     TEXT,
        price_usd_kg REAL,
        vol_tons     REAL,   -- MTD cumulative tons from SECEX weekly bulletin
        rev_000usd   REAL,   -- MTD cumulative revenue (thousands USD)
        biz_days     INTEGER -- business days reported in this bulletin period
    );
    CREATE TABLE IF NOT EXISTS monthly (
        period       TEXT PRIMARY KEY,
        year         INTEGER,
        month        INTEGER,
        secex_usd_kg REAL,
        fx           REAL,
        secex_brl_kg REAL,
        cepea_r_kg   REAL,
        spread       REAL,
        updated_at   TEXT
    );
    CREATE TABLE IF NOT EXISTS weekly (
        start_date   TEXT    PRIMARY KEY,
        end_date     TEXT,
        secex_usd_kg REAL,
        fx           REAL,
        secex_brl_kg REAL,
        cepea_r_kg   REAL,
        spread       REAL,
        vol_tons     REAL,        -- incremental weekly tons (de-accumulated from MTD)
        biz_days     INTEGER,     -- business days in this week's period
        vol_tons_daily REAL,      -- daily average = vol_tons / biz_days
        updated_at   TEXT
    );
    """)
    conn.commit()
    # ── Migrate existing DBs that pre-date these columns ─────────────────────
    migrations = [
        ("_weekly_raw", "vol_tons",   "REAL"),
        ("_weekly_raw", "rev_000usd", "REAL"),
        ("_weekly_raw", "biz_days",   "INTEGER"),
        ("weekly",      "vol_tons",   "REAL"),
        ("weekly",      "biz_days",   "INTEGER"),
        ("weekly",      "vol_tons_daily", "REAL"),
    ]
    for tbl, col, dtype in migrations:
        try:
            conn.execute(f"ALTER TABLE {tbl} ADD COLUMN {col} {dtype}")
            conn.commit()
            print(f"  [DB] Migrated: added {tbl}.{col}")
        except Exception:
            pass  # column already exists


# ══════════════════════════════════════════════════════════════════════════════
# BRAZIL BUSINESS-DAY HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def _easter(year):
    """Easter Sunday via Anonymous Gregorian algorithm."""
    a = year % 19; b = year // 100; c = year % 100
    d = b // 4;    e = b % 4;       f = (b + 8) // 25
    g = (b - f + 1) // 3;  h = (19*a + b - d - g + 15) % 30
    i = c // 4;    k = c % 4;       l = (32 + 2*e + 2*i - h - k) % 7
    m = (a + 11*h + 22*l) // 451
    mo  = (h + l - 7*m + 114) // 31
    day = ((h + l - 7*m + 114) % 31) + 1
    return date(year, mo, day)


def _br_holidays(year):
    """Set of ISO date strings for Brazilian national holidays in a given year."""
    from datetime import timedelta
    e = _easter(year)
    h = {
        str(date(year, 1,  1)),          # New Year
        str(e - timedelta(days=48)),     # Carnival Monday
        str(e - timedelta(days=47)),     # Carnival Tuesday
        str(e - timedelta(days=2)),      # Good Friday
        str(date(year, 4,  21)),         # Tiradentes
        str(date(year, 5,  1)),          # Labour Day
        str(date(year, 9,  7)),          # Independence Day
        str(date(year, 10, 12)),         # Nossa Senhora Aparecida
        str(date(year, 11, 2)),          # Finados
        str(date(year, 11, 15)),         # Proclamação da República
        str(date(year, 12, 25)),         # Christmas
    }
    if year >= 2024:                     # Consciência Negra became national in 2024
        h.add(str(date(year, 11, 20)))
    return h


def _biz_days_between(start_dt, end_dt):
    """Count Mon–Fri days (excl. BR national holidays) between dates, inclusive."""
    from datetime import timedelta
    hols = _br_holidays(start_dt.year)
    if end_dt.year != start_dt.year:
        hols |= _br_holidays(end_dt.year)
    count = 0
    d = start_dt
    while d <= end_dt:
        if d.weekday() < 5 and str(d) not in hols:
            count += 1
        d += timedelta(days=1)
    return count


def _nth_biz_day(year, month, n):
    """Return the date of the nth business day in month (1-indexed)."""
    from datetime import timedelta
    hols = _br_holidays(year)
    d = date(year, month, 1)
    count = 0
    for _ in range(60):   # safety bound
        if d.weekday() < 5 and str(d) not in hols:
            count += 1
            if count == n:
                return d
        d += timedelta(days=1)
    return None


# ══════════════════════════════════════════════════════════════════════════════
# HTTP HELPER
# ══════════════════════════════════════════════════════════════════════════════
def get(url, **kwargs):
    hdrs = {"Accept": "application/json", "User-Agent": "Mozilla/5.0"}
    # Brazilian government servers use ICP-Brasil certs not trusted on Linux by default
    ssl_verify = not any(h in url for h in _NO_VERIFY_HOSTS)
    for attempt in range(RETRY):
        try:
            r = requests.get(url, headers=hdrs, timeout=TIMEOUT,
                             verify=ssl_verify, **kwargs)
            r.raise_for_status()
            return r
        except Exception as e:
            if attempt == RETRY - 1:
                print(f"  ✗ {url[:70]}…: {e}")
                return None
            time.sleep(2 ** attempt)
    return None


# ══════════════════════════════════════════════════════════════════════════════
# BCB PTAX FX
# ══════════════════════════════════════════════════════════════════════════════
def fetch_fx(conn):
    """Download BCB PTAX BRL/USD daily rates into _fx_raw."""
    start = datetime(ANO_INI, 1, 1).strftime("%m-%d-%Y")
    end   = datetime.now().strftime("%m-%d-%Y")
    rows  = []

    # Method A: BCB OLINDA
    url = (
        "https://olinda.bcb.gov.br/olinda/servico/PTAX/versao/v1/odata/"
        f"CotacaoDolarPeriodo(dataInicial=@dataInicial,dataFinalCotacao=@dataFinalCotacao)"
        f"?@dataInicial='{start}'&@dataFinalCotacao='{end}'"
        "&$format=json&$select=cotacaoCompra,dataHoraCotacao"
    )
    r = get(url)
    if r:
        for item in r.json().get("value", []):
            try:
                rows.append((item["dataHoraCotacao"][:10], float(item["cotacaoCompra"])))
            except Exception:
                pass
        print(f"  [FX] BCB OLINDA: {len(rows)} rows")

    # Method B: BCB SGS series 1
    if not rows:
        url2 = (
            f"https://api.bcb.gov.br/dados/serie/bcdata.sgs.1/dados"
            f"?formato=json&dataInicial=01/01/{ANO_INI}&dataFinal="
            + datetime.now().strftime("%d/%m/%Y")
        )
        r2 = get(url2)
        if r2:
            for item in r2.json():
                try:
                    dt = datetime.strptime(item["data"], "%d/%m/%Y").strftime("%Y-%m-%d")
                    rows.append((dt, float(item["valor"])))
                except Exception:
                    pass
            print(f"  [FX] BCB SGS: {len(rows)} rows")

    if not rows:
        print("  [FX] All methods failed — FX not updated.")
        return 0

    conn.executemany("INSERT OR REPLACE INTO _fx_raw(dt,fx) VALUES(?,?)", rows)
    conn.commit()
    print(f"  [FX] {len(rows)} rows stored.")
    return len(rows)


# ══════════════════════════════════════════════════════════════════════════════
# SECEX MONTHLY
# ══════════════════════════════════════════════════════════════════════════════
def fetch_secex(conn, years=None):
    """Download MDIC annual CSVs and upsert into _secex_raw."""
    from io import StringIO
    try:
        import pandas as pd
    except ImportError:
        sys.exit("Missing: pip install pandas openpyxl")

    if years is None:
        years = range(ANO_INI, datetime.now().year + 1)

    # The annual CSV uses 8-digit NCM codes (e.g. "02011000").
    # NCM_CODES contains 4-digit chapter codes ("0201", "0202"), so we
    # match by prefix — str[:4] — rather than exact equality.
    BASE = "https://balanca.economia.gov.br/balanca/bd/comexstat-bd/ncm/EXP_{year}.csv"
    total = 0
    for yr in years:
        r = get(BASE.format(year=yr))
        if not r:
            continue
        try:
            df = pd.read_csv(StringIO(r.text), sep=";", dtype=str, low_memory=False)
            # Normalise: strip whitespace and zero-pad to 8 chars
            df["CO_NCM"] = df["CO_NCM"].str.strip().str.zfill(8)
            # Keep rows whose 4-digit chapter prefix matches NCM_CODES
            df = df[df["CO_NCM"].str[:4].isin(NCM_CODES)].copy()
            if df.empty:
                print(f"  [SECEX] {yr}: no beef rows")
                continue
            df["CO_MES"]     = df["CO_MES"].astype(int)
            df["KG_LIQUIDO"] = df["KG_LIQUIDO"].astype(float)
            df["VL_FOB"]     = df["VL_FOB"].astype(float)
            grp = df.groupby("CO_MES").agg(
                vol_kg=("KG_LIQUIDO","sum"), rev_usd=("VL_FOB","sum")
            ).reset_index()
            rows = []
            for _, row in grp.iterrows():
                m    = int(row["CO_MES"])
                vol  = float(row["vol_kg"]) / 1000      # tons
                rev  = float(row["rev_usd"]) / 1000     # 000 USD
                p    = (rev * 1000 / (vol * 1000)) if vol > 0 else None  # USD/kg
                rows.append((yr, m, rev, vol, p))
            conn.executemany(
                "INSERT OR REPLACE INTO _secex_raw(year,month,rev_000usd,vol_tons,price_usd_kg)"
                " VALUES(?,?,?,?,?)", rows
            )
            conn.commit()
            total += len(rows)
            print(f"  [SECEX] {yr}: {len(rows)} months")
        except Exception as ex:
            print(f"  [SECEX] {yr}: {ex}")
    return total


# ══════════════════════════════════════════════════════════════════════════════
# CEPEA (local XLS — seeding only)
# ══════════════════════════════════════════════════════════════════════════════
def load_cepea_xls(conn, path):
    """Load CEPEA Boi Gordo R$/arroba from local XLS/XLSX file."""
    import subprocess, tempfile
    try:
        import pandas as pd
    except ImportError:
        sys.exit("Missing: pip install pandas openpyxl")

    p = Path(path)
    if not p.exists():
        print(f"  [CEPEA] Not found: {path}")
        return 0

    df = None
    # Try direct read (XLSX)
    try:
        df = pd.read_excel(p, header=None, engine="openpyxl")
    except Exception:
        pass

    # LibreOffice fallback (OLE2 .xls)
    if df is None:
        try:
            tmp = tempfile.mkdtemp()
            out = Path(tmp) / (p.stem + "_conv.xlsx")
            subprocess.run(
                ["libreoffice", "--headless", "--convert-to", "xlsx", str(p), "--outdir", tmp],
                capture_output=True, timeout=90
            )
            if out.exists():
                df = pd.read_excel(out, header=None, engine="openpyxl")
                print("  [CEPEA] LibreOffice conversion OK")
        except Exception as ex:
            print(f"  [CEPEA] LibreOffice failed: {ex}")

    if df is None:
        print("  [CEPEA] Could not read file.")
        return 0

    rows = []
    for _, row in df.iterrows():
        for c in range(len(row) - 1):
            cell = row.iloc[c]
            dt   = None
            if isinstance(cell, (datetime, date)):
                dt = cell.date() if isinstance(cell, datetime) else cell
            elif isinstance(cell, str):
                for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
                    try:
                        dt = datetime.strptime(cell.strip(), fmt).date()
                        break
                    except Exception:
                        pass
            if dt is None or dt.year < 2000:
                continue
            try:
                val = float(str(row.iloc[c + 1]).replace(",", "."))
                if 50 < val < 2000:
                    rows.append((str(dt), val, round(val / 15.0, 6)))
            except Exception:
                pass

    if not rows:
        print("  [CEPEA] No valid rows parsed.")
        return 0

    conn.executemany(
        "INSERT OR REPLACE INTO _cepea_raw(dt,r_arroba,r_kg) VALUES(?,?,?)", rows
    )
    conn.commit()
    print(f"  [CEPEA] {len(rows)} rows loaded from {p.name}")
    return len(rows)


# ══════════════════════════════════════════════════════════════════════════════
# WEEKLY RAW SEED
# ══════════════════════════════════════════════════════════════════════════════
def _enrich_seed():
    """
    Add biz_days to each WEEKLY_SEED row.

    WEEKLY_SEED already carries exact rev_000usd_mtd values sourced directly
    from the SECEX bulletin (no rounded-price arithmetic), so this function
    only needs to compute business-day counts.

    Returns list of (start, end, price, vol_mtd, rev_mtd, biz_days).
    """
    result = []
    for row in WEEKLY_SEED:
        s, e, price, vol_mtd, rev_mtd = row
        bd = _biz_days_between(date.fromisoformat(s), date.fromisoformat(e))
        result.append((s, e, price, vol_mtd, rev_mtd, bd))
    return result


def seed_weekly_raw(conn):
    """Seed _weekly_raw from WEEKLY_SEED with smart conflict resolution.

    Computes and stores MTD-cumulative revenue (rev_000usd) and biz_days for
    each row, enabling incremental price de-accumulation in materialise().
    """
    enriched = _enrich_seed()
    conn.executemany(
        """
        INSERT INTO _weekly_raw(start_date, end_date, price_usd_kg,
                                vol_tons, rev_000usd, biz_days)
        VALUES (?, ?, ?, ?, ?, ?)
        ON CONFLICT(start_date) DO UPDATE SET
            -- SEED is always authoritative: fully replace all fields.
            -- Partial updates (the old approach) caused price corruption when
            -- rev was updated but vol/end_date were left from a stale bulletin row.
            end_date     = excluded.end_date,
            price_usd_kg = excluded.price_usd_kg,
            vol_tons     = excluded.vol_tons,
            rev_000usd   = excluded.rev_000usd,
            biz_days     = excluded.biz_days
        """,
        enriched
    )
    conn.commit()
    print(f"  [WEEKLY] {len(WEEKLY_SEED)} rows seeded/validated in _weekly_raw.")


# ══════════════════════════════════════════════════════════════════════════════
# CEPEA WEB SCRAPER (incremental daily updates)
# ══════════════════════════════════════════════════════════════════════════════
def fetch_cepea_web(conn):
    """
    Fetch recent CEPEA Boi Gordo R$/arroba indicator data from the web and
    insert any dates newer than what is already in _cepea_raw.

    Strategy (in order):
      1. CEPEA widget endpoint (/widgetpec/cotacao.aspx) — returns static HTML
         with a table; try multiple known indicator IDs for boi gordo
      2. Main indicator page → follow the "SÉRIE DE PREÇOS" XLS download link
      3. Try direct XLS URL patterns as last resort

    NOTE: The main CEPEA indicator page (boi-gordo.aspx) is a React SPA — its
    table is JS-rendered and NOT in the static HTML. The widget endpoint is
    static and machine-readable.

    CEPEA boi gordo indicator IDs (try all; the correct one may vary):
      indicador=2  (boi gordo carcaça — most common)
      indicador=1  (alternate)
      indicador=3  (alternate)
    """
    import io, re as _re, json as _json
    BASE = "https://www.cepea.org.br"
    PAGE_URL = f"{BASE}/br/indicador/boi-gordo.aspx"

    last_dt = conn.execute("SELECT MAX(dt) FROM _cepea_raw").fetchone()[0]
    print(f"  [CEPEA-WEB] Last date in DB: {last_dt}")

    try:
        import requests as _req
    except ImportError:
        print("  [CEPEA-WEB] Missing: pip install requests")
        return 0

    HDRS = {"User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "pt-BR,pt;q=0.9,en;q=0.8"}

    parsed_rows = []

    # ─────────────────────────────────────────────────────────────────────────
    # Helper: extract (iso_date, r_arroba) pairs from an HTML string
    # Works on CEPEA widget HTML which has tables like:
    #   <td>17/04/2026</td><td>365,10</td>...
    # ─────────────────────────────────────────────────────────────────────────
    def _rows_from_html(html):
        rows = []
        # Strip tags inside <td> to get clean text values
        tds = _re.findall(r'<td[^>]*>\s*(.*?)\s*</td>', html, _re.IGNORECASE | _re.DOTALL)
        # Walk pairs: find a date cell, then grab the next numeric cell
        i = 0
        while i < len(tds):
            raw = _re.sub(r'<[^>]+>', '', tds[i]).strip()
            dt = None
            for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                try:
                    dt = datetime.strptime(raw, fmt).date(); break
                except Exception:
                    pass
            if dt and dt.year >= 2005:
                # Look ahead up to 3 cells for a numeric value
                for j in range(i + 1, min(i + 4, len(tds))):
                    num_raw = _re.sub(r'<[^>]+>', '', tds[j]).strip()
                    num_raw = num_raw.replace(".", "").replace(",", ".").replace("R$", "").strip()
                    try:
                        val = float(num_raw)
                        if 150 < val < 3000:   # R$/arroba sanity range
                            rows.append((str(dt), val, round(val / 15.0, 6)))
                            break
                    except Exception:
                        pass
            i += 1
        return rows

    # ─────────────────────────────────────────────────────────────────────────
    # Helper: parse XLS/XLSX bytes
    # ─────────────────────────────────────────────────────────────────────────
    def _rows_from_xls(content):
        import io as _io
        rows = []
        try:
            import openpyxl as _xl
            wb = _xl.load_workbook(_io.BytesIO(content), data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                if not row:
                    continue
                for ci in range(len(row) - 1):
                    cell = row[ci]
                    dt = None
                    if isinstance(cell, datetime):
                        dt = cell.date()
                    elif isinstance(cell, date) and not isinstance(cell, datetime):
                        dt = cell
                    elif isinstance(cell, str):
                        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                            try:
                                dt = datetime.strptime(cell.strip(), fmt).date(); break
                            except Exception:
                                pass
                    if dt is None or dt.year < 2005:
                        continue
                    try:
                        val = float(str(row[ci + 1]).replace(",", ".").replace("R$", "").strip())
                        if 150 < val < 3000:
                            rows.append((str(dt), val, round(val / 15.0, 6)))
                    except Exception:
                        pass
        except Exception as exc:
            print(f"  [CEPEA-WEB] XLS parse error: {exc}")
        return rows

    # ── Step 0: NoticiasAgricolas — mirrors CEPEA/ESALQ, non-CEPEA IP ────────
    # This site republishes the official CEPEA/ESALQ boi gordo indicator and
    # is not hosted on CEPEA infrastructure, so Azure IP blocks don't apply.
    # The cotacoes page returns a table with recent daily prices.
    NA_URLS = [
        "https://www.noticiasagricolas.com.br/cotacoes/boi-gordo/boi-gordo-indicador-esalq-bmf",
        "https://www.noticiasagricolas.com.br/widgets/cotacoes?id=12",
    ]
    for na_url in NA_URLS:
        try:
            r0 = _req.get(na_url, headers=HDRS, timeout=15, verify=False)
            if r0.status_code == 200:
                rows = _rows_from_html(r0.text)
                if rows:
                    parsed_rows = rows
                    print(f"  [CEPEA-WEB] NoticiasAgricolas ({na_url.split('/')[-1][:30]}): {len(rows)} rows")
                    break
                print(f"  [CEPEA-WEB] NoticiasAgricolas: HTTP 200 but no rows parsed")
            else:
                print(f"  [CEPEA-WEB] NoticiasAgricolas: HTTP {r0.status_code}")
        except Exception as exc:
            print(f"  [CEPEA-WEB] NoticiasAgricolas: {exc}")

    # ── Step 1: CEPEA widget endpoint (static HTML, not React) ───────────────
    # The widget page serves a simple HTML table without JavaScript rendering.
    # Boi gordo carcaça = indicador 2 (try multiple IDs for robustness).
    widget_ids = [2, 1, 3, 28, 17]
    for ind_id in widget_ids:
        widget_url = f"{BASE}/br/widgetpec/cotacao.aspx?indicador={ind_id}"
        try:
            wr = _req.get(widget_url, headers=HDRS, timeout=30, verify=False)
            if wr.status_code != 200:
                print(f"  [CEPEA-WEB] Widget id={ind_id}: HTTP {wr.status_code}")
                continue
            html = wr.text
            # Confirm this is the boi gordo widget (page title or header should say "boi")
            if "boi" not in html.lower() and "bovino" not in html.lower():
                print(f"  [CEPEA-WEB] Widget id={ind_id}: not boi gordo page — skipping")
                continue
            rows = _rows_from_html(html)
            if rows:
                parsed_rows = rows
                print(f"  [CEPEA-WEB] Widget id={ind_id}: {len(rows)} rows parsed")
                break
            print(f"  [CEPEA-WEB] Widget id={ind_id}: fetched but no rows in table")
        except Exception as exc:
            print(f"  [CEPEA-WEB] Widget id={ind_id}: {exc}")

    # ── Step 2: main indicator page → find XLS download link ─────────────────
    if not parsed_rows:
        page_html = ""
        try:
            pr = _req.get(PAGE_URL, headers=HDRS, timeout=30, verify=False)
            page_html = pr.text
            print(f"  [CEPEA-WEB] Indicator page: {len(page_html):,} chars")
        except Exception as exc:
            print(f"  [CEPEA-WEB] Indicator page error: {exc}")

        if page_html:
            # 2a: Any XLS/XLSX link in the page (href or JS string)
            xls_links = list(dict.fromkeys(
                _re.findall(r'["\']([^"\']*\.xls[x]?)["\']', page_html, _re.IGNORECASE)
            ))
            for lnk in xls_links[:10]:
                url = lnk if lnk.startswith("http") else BASE + ("" if lnk.startswith("/") else "/") + lnk
                try:
                    rx = _req.get(url, headers=HDRS, timeout=40, verify=False)
                    if rx.status_code == 200 and len(rx.content) > 1000:
                        rows = _rows_from_xls(rx.content)
                        if rows:
                            parsed_rows = rows
                            print(f"  [CEPEA-WEB] XLS from page link ({lnk[-40:]}): {len(rows)} rows")
                            break
                except Exception as exc:
                    print(f"  [CEPEA-WEB] XLS link error ({lnk[-30:]}): {exc}")

            # 2b: Try JSON blobs in <script> tags (React initial state)
            if not parsed_rows:
                for sc in _re.findall(r'<script[^>]*>(.*?)</script>', page_html, _re.DOTALL):
                    # Look for arrays containing date-like strings and prices
                    for arr_str in _re.findall(r'\[[^\[\]]{50,}\]', sc):
                        try:
                            import json as _js
                            arr = _js.loads(arr_str)
                            for item in arr:
                                if not isinstance(item, dict):
                                    continue
                                for dk in ("data", "date", "Data", "dt", "DATA"):
                                    for vk in ("preco", "valor", "price", "Preco", "cotacao", "r_arroba"):
                                        dv = item.get(dk); vv = item.get(vk)
                                        if dv and vv:
                                            try:
                                                for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                                                    try:
                                                        dt = datetime.strptime(str(dv).strip(), fmt).date(); break
                                                    except Exception:
                                                        dt = None
                                                val = float(str(vv).replace(",", "."))
                                                if dt and dt.year >= 2005 and 150 < val < 3000:
                                                    parsed_rows.append((str(dt), val, round(val/15.0, 6)))
                                            except Exception:
                                                pass
                        except Exception:
                            pass
                if parsed_rows:
                    print(f"  [CEPEA-WEB] {len(parsed_rows)} rows from JSON in page scripts")

    # ── Step 3: direct XLS URL patterns (last resort) ─────────────────────────
    if not parsed_rows:
        direct_urls = [
            f"{BASE}/br/uploads/indicador/boi.xls",
            f"{BASE}/br/uploads/indicador/boi-gordo.xls",
            f"{BASE}/br/uploads/indicador/boi.xlsx",
            f"{BASE}/br/uploads/boi.xls",
            f"{BASE}/uploads/indicador/boi.xls",
        ]
        for url in direct_urls:
            try:
                rx = _req.get(url, headers=HDRS, timeout=40, verify=False)
                if rx.status_code == 200 and len(rx.content) > 1000:
                    rows = _rows_from_xls(rx.content)
                    if rows:
                        parsed_rows = rows
                        print(f"  [CEPEA-WEB] Direct XLS {url.split('/')[-1]}: {len(rows)} rows")
                        break
                    print(f"  [CEPEA-WEB] {url.split('/')[-1]}: HTTP 200 but no rows")
                else:
                    print(f"  [CEPEA-WEB] {url.split('/')[-1]}: HTTP {rx.status_code}")
            except Exception as exc:
                print(f"  [CEPEA-WEB] {url.split('/')[-1]}: {exc}")

    if not parsed_rows:
        print("  [CEPEA-WEB] All strategies failed — CEPEA not updated.")
        print("  [CEPEA-WEB] Download the XLS manually from:")
        print("  [CEPEA-WEB]   https://www.cepea.org.br/br/indicador/boi-gordo.aspx")
        print("  [CEPEA-WEB]   (click 'SÉRIE DE PREÇOS' → download XLS)")
        print("  [CEPEA-WEB] Then seed: python extractor_bz.py --init --cepea /path/to/file.xls")
        return 0

    # ── De-duplicate and persist only new rows ────────────────────────────────
    seen = set()
    unique_rows = []
    for row in parsed_rows:
        if row[0] not in seen:
            seen.add(row[0])
            unique_rows.append(row)
    unique_rows.sort(key=lambda r: r[0])

    new_rows = [(dt, ra, rk) for dt, ra, rk in unique_rows
                if last_dt is None or dt > last_dt]
    if not new_rows:
        print(f"  [CEPEA-WEB] Already up-to-date (latest in DB: {last_dt})")
        return 0

    conn.executemany(
        "INSERT OR REPLACE INTO _cepea_raw(dt, r_arroba, r_kg) VALUES(?,?,?)",
        new_rows
    )
    conn.commit()
    latest = max(dt for dt, _, _ in new_rows)
    print(f"  [CEPEA-WEB] Inserted {len(new_rows)} new rows (up to {latest})")
    return len(new_rows)


# ══════════════════════════════════════════════════════════════════════════════
# FETCH WEEKLY SECEX BULLETIN (price + MTD volume)
# ══════════════════════════════════════════════════════════════════════════════
def fetch_weekly_bulletin(conn):
    """
    Fetch the SECEX/MDIC weekly bulletin Excel ("Produto por Atividade Econômica"
    — CUCI classification) from balanca.economia.gov.br and update _weekly_raw
    with the latest week's price_usd_kg and vol_tons (MTD cumulative).

    The bulletin page is a React SPA — links are rendered via JavaScript and
    won't appear in raw HTML. This function uses three discovery strategies:
      1. Search the raw HTML for any xlsx URL (href attrs AND JS strings)
      2. Try a set of guessed URL patterns based on known SECEX file naming
      3. As a last resort, try the SECEX API endpoint directly

    Excel structure (per current month):
      US$ Mil | US$ Mil/avg | Toneladas | Toneladas/avg | Preço (US$/Ton) | Var%

    We extract for "Carne bovina fresca, refrigerada ou congelada":
      - US$ Mil   → revenue (MTD, thousands USD)
      - Toneladas → volume (MTD, tons)

    De-accumulation (MTD → weekly) happens later in materialise().
    """
    import io, re as _re

    BASE = "https://balanca.economia.gov.br"
    PAGE_URL = f"{BASE}/balanca/pg_principal_bc/principais_resultados.html"

    from datetime import date as _date, timedelta as _td
    PT_MON = {1:"Jan",2:"Fev",3:"Mar",4:"Abr",5:"Mai",6:"Jun",
              7:"Jul",8:"Ago",9:"Set",10:"Out",11:"Nov",12:"Dez"}
    PT_MON_REV = {v.lower(): k for k, v in PT_MON.items()}
    today  = _date.today()
    # yr/mo will be overridden once we parse the bulletin period from the Excel
    yr     = today.year
    mo     = today.month
    wk_num = today.isocalendar()[1]

    # ── Step 1: fetch HTML and search broadly for xlsx links ──────────────────
    # NOTE: we use requests directly (not get()) to avoid the duplicate-headers
    #       bug that occurs when get()'s hdrs= and **kwargs both carry 'headers'.
    xlsx_url = None
    KEYWORDS = ("cuci", "produto", "semana", "boletim", "isic", "ativ")
    try:
        import requests as _req
        html_r = _req.get(
            PAGE_URL,
            headers={"Accept": "text/html,application/xhtml+xml",
                     "User-Agent": "Mozilla/5.0"},
            timeout=30,
            verify=False,
        )
        page_text = html_r.text

        # Search 1a: href attributes containing .xls / .xlsx
        href_links = _re.findall(r'href=["\']([^"\']+\.xlsx?)["\']',
                                 page_text, _re.IGNORECASE)
        # Search 1b: any URL-like string with .xlsx in entire page source
        #            (catches JS bundle strings like "/path/file.xlsx")
        all_xlsx = _re.findall(r'["\']([^"\']*\.xlsx?)["\']',
                                page_text, _re.IGNORECASE)

        candidates = href_links + all_xlsx
        print(f"  [BULLETIN] Page fetched ({len(page_text):,} chars). "
              f"xlsx candidates found: {len(candidates)}")

        for lnk in candidates:
            if any(k in lnk.lower() for k in KEYWORDS):
                xlsx_url = lnk if lnk.startswith("http") else f"{BASE}{lnk}"
                print(f"  [BULLETIN] Found via page scrape: {xlsx_url}")
                break

        if not xlsx_url and candidates:
            print(f"  [BULLETIN] Candidates (no keyword match): {candidates[:8]}")

    except Exception as exc:
        print(f"  [BULLETIN] Page fetch error: {exc}")

    # ── Step 2: guessed URL patterns (SECEX naming conventions) ──────────────
    if xlsx_url is None:
        print("  [BULLETIN] Trying guessed URL patterns …")
        guesses = []
        for w in range(wk_num, wk_num - 3, -1):    # current week and 2 prior
            w = max(w, 1)
            for tpl in (
                f"/balanca/bd/boletim/CUCI_EXP_SEMANA_{yr}_{w:02d}.xlsx",
                f"/balanca/bd/boletim/AtividadeEconomica_EXP_SEMANA_{yr}_{w:02d}.xlsx",
                f"/balanca/bd/boletim/PRODUTO_EXP_SEMANA_{yr}_{w:02d}.xlsx",
                f"/balanca/bd/boletim/CUCI_EXP_SEMANA_{yr}_{mo:02d}.xlsx",
                f"/balanca/bd/boletim/CUCI_EXP_SEMANA_{yr}_{mo:02d}_{today.day:02d}.xlsx",
            ):
                guesses.append(BASE + tpl)

        for url in guesses:
            try:
                import requests as _req
                probe = _req.head(url, timeout=10, verify=False,
                                  headers={"User-Agent": "Mozilla/5.0"})
                if probe.status_code == 200:
                    xlsx_url = url
                    print(f"  [BULLETIN] Guessed URL found: {xlsx_url}")
                    break
                else:
                    print(f"  [BULLETIN]   {probe.status_code} {url.split('/')[-1]}")
            except Exception:
                pass

    if xlsx_url is None:
        print("  [BULLETIN] Could not locate Excel file. "
              "Set env var BULLETIN_XLSX_URL to override, e.g.:")
        print("  export BULLETIN_XLSX_URL='https://balanca.economia.gov.br/balanca/bd/boletim/CUCI_EXP_SEMANA_XXXX_YY.xlsx'")
        # Try env var override as last resort
        import os
        xlsx_url = os.environ.get("BULLETIN_XLSX_URL")
        if xlsx_url:
            print(f"  [BULLETIN] Using env override: {xlsx_url}")
        else:
            return 0

    # ── Step 3: download Excel ────────────────────────────────────────────────
    print(f"  [BULLETIN] Downloading: {xlsx_url}")
    r2 = get(xlsx_url)
    if r2 is None:
        return 0

    try:
        import openpyxl, zipfile as _zf, re as _re2
        raw_bytes = r2.content
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
        except Exception as exc1:
            # openpyxl bug: Excel has broken drawing references in the zip.
            # Fix: rebuild zip without drawing files and strip drawing rels.
            if "drawing" in str(exc1).lower() or "no item named" in str(exc1).lower():
                print(f"  [BULLETIN] Drawing ref error — stripping drawings and retrying …")
                buf_fix = io.BytesIO()
                with _zf.ZipFile(io.BytesIO(raw_bytes), 'r') as zin:
                    with _zf.ZipFile(buf_fix, 'w', _zf.ZIP_DEFLATED) as zout:
                        for item in zin.infolist():
                            if 'drawing' in item.filename.lower():
                                continue          # drop drawing files
                            data = zin.read(item.filename)
                            if item.filename.endswith('.rels'):
                                # Remove <Relationship> entries pointing to drawings
                                data = _re2.sub(
                                    rb'<Relationship[^>]+/drawing[^>]+/?>',
                                    b'', data
                                )
                            zout.writestr(item, data)
                buf_fix.seek(0)
                wb = openpyxl.load_workbook(buf_fix, data_only=True)
            else:
                print(f"  [BULLETIN] Excel parse error: {exc1}")
                return 0
    except Exception as exc:
        print(f"  [BULLETIN] Excel parse error: {exc}")
        return 0

    # ── Step 4: find the right sheet and parse headers ────────────────────────
    # The workbook may have multiple sheets; try the active one first, then all
    sheets_to_try = [wb.active] + [wb[s] for s in wb.sheetnames if wb[s] != wb.active]

    # The Setores_Produtos.xlsx structure (confirmed from debug):
    #   Header row:  [Descrição | US$ Mil | US$ Mil/MédDiária | Toneladas |
    #                 Ton/MédDiária | Preço (US$/Tonelada) | Variação (%)]
    #   Each category has 2 data cols: [current period | previous period]
    #   "Current period" is always the FIRST column after each category header.
    #   There are NO per-column period sub-headers; the period label appears
    #   only in the "Variação" column as "MonAno - MonAno".
    #   → We find category header columns DIRECTLY and use them as data columns.

    def _parse_sheet(ws):
        """
        Return (vol_tons_mtd, price_usd_kg, rev_000usd_mtd, biz_days,
                bull_year, bull_month).

        Excel layout (Setores_Produtos.xlsx):
          Descrição | US$ Mil | US$ Mil/MédDiária | Toneladas |
          Ton/MédDiária | Preço (US$/Tonelada) | Variação (%)

        We extract for "Carne bovina fresca, refrigerada ou congelada":
          • Toneladas          → vol_tons_mtd  (MTD cumulative)
          • US$ Mil            → rev_000usd_mtd (MTD cumulative revenue)
          • Ton/MédDiária      → ton_media_diaria → biz_days = round(vol/media)
          • Preço (US$/Ton)    → stored as reference only; we prefer
                                  de-accumulated price in materialise()

        Period (month/year) is extracted from a header cell matching
        Portuguese month abbreviation + 4-digit year.
        """
        # ── Extract bulletin period from header rows ──────────────────────────
        bull_year = bull_month = None
        for hrow in ws.iter_rows(max_row=10):
            for cell in hrow:
                v = str(cell.value or "").strip()
                m = _re.search(
                    r'\b(jan|fev|mar|abr|mai|jun|jul|ago|set|out|nov|dez)[a-z]*/(\d{4})\b',
                    v, _re.IGNORECASE)
                if m:
                    bull_month = PT_MON_REV[m.group(1).lower()[:3]]
                    bull_year  = int(m.group(2))
                    break
            if bull_year:
                break

        ton_col        = None   # 1-based col: "Toneladas"
        ton_media_col  = None   # 1-based col: "Ton/MédDiária"  (next col after ton_col)
        price_col      = None   # 1-based col: "Preço (US$/Ton)"
        usd_col        = None   # 1-based col: "US$ Mil"
        cat_header_row = None

        for row in ws.iter_rows(max_row=30):
            for cell in row:
                v   = str(cell.value or "").strip().lower()
                col = cell.column
                if v.startswith("ton") and "media" not in v and "med" not in v \
                        and ton_col is None:
                    ton_col        = col
                    # Excel has curr/prev pairs per metric:
                    # Ton_curr | Ton_prev | Ton/MédDiária_curr | ...
                    # So Ton/MédDiária is 2 columns after Ton (not 1)
                    ton_media_col  = col + 2
                    cat_header_row = cell.row
                if ("preço" in v or "preco" in v or "us$/ton" in v) and price_col is None:
                    price_col = col
                if "us$ mil" in v and "media" not in v and "med" not in v \
                        and usd_col is None:
                    usd_col = col
            if ton_col is not None and cat_header_row is not None:
                if price_col is not None or usd_col is not None:
                    break

        if ton_col is None or cat_header_row is None:
            return None, None, None, None, None, None

        # ── Find "Carne bovina fresca" row ────────────────────────────────────
        carne_row = None
        for row in ws.iter_rows(min_row=cat_header_row + 1):
            for cell in row[:3]:
                v = str(cell.value or "").lower()
                if "carne bovina" in v and "fresca" in v:
                    carne_row = cell.row
                    break
            if carne_row:
                break

        if carne_row is None:
            return None, None, None, None, None, None

        # ── Extract values ────────────────────────────────────────────────────
        vol_raw      = ws.cell(row=carne_row, column=ton_col).value
        vol_tons     = float(vol_raw) if vol_raw is not None else None

        # Revenue (MTD, thousands USD)
        rev_raw      = ws.cell(row=carne_row, column=usd_col).value if usd_col else None
        rev_000usd   = float(rev_raw) if rev_raw is not None else None

        # Business days from Ton/MédDiária: biz_days = round(vol_MTD / daily_avg)
        biz_days = None
        if ton_media_col and vol_tons:
            media_raw = ws.cell(row=carne_row, column=ton_media_col).value
            if media_raw:
                try:
                    media_val = float(media_raw)
                    if media_val > 0:
                        biz_days = round(vol_tons / media_val)
                except Exception:
                    pass

        # Price from "Preço" column — stored for reference / sanity check only
        if price_col is not None:
            price_per_ton = ws.cell(row=carne_row, column=price_col).value
            price_usd_kg  = (float(price_per_ton) / 1000.0) if price_per_ton else None
        elif rev_000usd is not None and vol_tons:
            price_usd_kg  = rev_000usd / vol_tons   # MTD cumulative price fallback
        else:
            price_usd_kg  = None

        print(f"  [BULLETIN] Cols → usd={usd_col}, ton={ton_col}, "
              f"ton_media={ton_media_col}, price={price_col} | beef row={carne_row}")
        print(f"  [BULLETIN] Extracted: vol_MTD={vol_tons!r}  "
              f"rev_MTD={rev_000usd!r}  biz_days={biz_days!r}")
        print(f"  [BULLETIN] Period from header: {bull_month}/{bull_year}")

        return vol_tons, price_usd_kg, rev_000usd, biz_days, bull_year, bull_month

    vol_mtd    = None
    price_usd  = None
    rev_mtd    = None
    biz_days   = None
    used_sheet = None
    bull_yr    = None
    bull_mo    = None
    for ws in sheets_to_try:
        v, p, rev, bd, by, bm = _parse_sheet(ws)
        if v is not None:
            vol_mtd    = v
            price_usd  = p
            rev_mtd    = rev
            biz_days   = bd
            bull_yr    = by
            bull_mo    = bm
            used_sheet = ws.title
            print(f"  [BULLETIN] Sheet '{used_sheet}': vol_MTD={vol_mtd:,.0f} t"
                  + (f", rev={rev_mtd:,.1f} 000 USD" if rev_mtd else "")
                  + (f", price={price_usd:.4f} USD/kg" if price_usd else "")
                  + (f", biz_days={biz_days}" if biz_days else ""))
            break

    if vol_mtd is None:
        print("  [BULLETIN] Could not parse Excel. Sheet headers dump:")
        for ws in sheets_to_try[:2]:
            print(f"  Sheet: '{ws.title}'")
            for i, row in enumerate(ws.iter_rows(max_row=10)):
                vals = [(c.column, str(c.value or "")[:25]) for c in row if c.value]
                if vals:
                    print(f"    row {i+1}: {vals}")
        return 0

    # ── Sanity check on price ─────────────────────────────────────────────────
    PRICE_MIN, PRICE_MAX = 3.0, 20.0
    if price_usd is not None and not (PRICE_MIN <= price_usd <= PRICE_MAX):
        print(f"  [BULLETIN] Price {price_usd:.4f} out of range [{PRICE_MIN},{PRICE_MAX}] — discarding.")
        price_usd = None

    # ── Step 5: resolve bulletin period and locate/create the target row ──────
    # Use the bulletin's own reported period (from Excel header), NOT today.
    # This prevents phantom rows when the bulletin still reports the previous month.
    target_yr = bull_yr if bull_yr else yr
    target_mo = bull_mo if bull_mo else mo
    yr_s = f"{target_yr:04d}"
    mo_s = f"{target_mo:02d}"
    print(f"  [BULLETIN] Bulletin period resolved to: {yr_s}-{mo_s}"
          + (f"  ({biz_days} biz days)" if biz_days else ""))

    existing = conn.execute(
        "SELECT start_date, end_date, price_usd_kg, biz_days FROM _weekly_raw"
        " WHERE start_date LIKE ? ORDER BY start_date DESC LIMIT 1",
        (f"{yr_s}-{mo_s}-%",)
    ).fetchone()

    if existing is None:
        # Only create a new row when the bulletin IS for the current calendar month
        if target_yr == today.year and target_mo == today.month:
            # Start: always the 1st of the month (MTD starts from day 1)
            s_date = f"{yr_s}-{mo_s}-01"
            # End: compute from biz_days if available, else use today
            if biz_days:
                end_dt = _nth_biz_day(target_yr, target_mo, biz_days)
                e_date = str(end_dt) if end_dt else str(today)
            else:
                e_date = str(today - _td(days=max(0, today.weekday() - 4)) if today.weekday() > 4 else today)
            existing_price = None
            print(f"  [BULLETIN] No existing row — creating new row {s_date} → {e_date}")
        else:
            print(f"  [BULLETIN] No row for {yr_s}-{mo_s} and it is a past month — skipping.")
            return 0
    else:
        s_date, e_date, existing_price, existing_bd = existing

        # Fix phantom rows: if start_date is not the 1st of the month
        # (e.g. "2026-04-13"), delete it and reset to "YYYY-MM-01"
        if not s_date.endswith("-01"):
            phantom = s_date
            s_date = f"{yr_s}-{mo_s}-01"
            conn.execute("DELETE FROM _weekly_raw WHERE start_date = ?", (phantom,))
            conn.commit()
            print(f"  [BULLETIN] Deleted phantom row {phantom} — will insert {s_date}")
            existing_price = None
            existing_bd    = None
            e_date = str(today)

        # ── NEW BULLETIN DETECTION ────────────────────────────────────────────
        # Compare the end_date the new bulletin would produce against the
        # latest stored row's end_date.  Using end_date (rather than biz_days)
        # avoids the MTD-vs-weekly confusion: SEED rows store weekly biz_days
        # while bulletin rows store MTD cumulative biz_days.
        new_end_dt = _nth_biz_day(target_yr, target_mo, biz_days) if biz_days else None
        new_e_str  = str(new_end_dt) if new_end_dt else str(today)

        if new_e_str > e_date:
            # Bulletin covers a period BEYOND the latest stored row → new row
            prev_end_dt  = date.fromisoformat(e_date)
            new_start_dt = prev_end_dt + _td(days=1)
            while new_start_dt.weekday() >= 5:   # skip weekends
                new_start_dt += _td(days=1)
            s_date = str(new_start_dt)
            e_date = new_e_str
            existing_price = None
            print(f"  [BULLETIN] New bulletin detected (end {e_date} → {new_e_str})"
                  f" — inserting NEW row {s_date} → {e_date}")
        else:
            # Same or older bulletin — update existing row in place
            if new_e_str != e_date and new_end_dt:
                print(f"  [BULLETIN] Correcting end_date: {e_date} → {new_e_str}")
                e_date = new_e_str
            print(f"  [BULLETIN] Updating existing row: {s_date} → {e_date}")

    # Validate existing_price; don't carry forward a previously bad value
    if existing_price is not None and not (PRICE_MIN <= existing_price <= PRICE_MAX):
        print(f"  [BULLETIN] Existing price {existing_price:.4f} also out of range — clearing.")
        existing_price = None

    # Store the MTD cumulative price as reference (real price computed in materialise)
    final_price = price_usd if price_usd is not None else existing_price

    conn.execute(
        "INSERT OR REPLACE INTO _weekly_raw"
        "(start_date, end_date, price_usd_kg, vol_tons, rev_000usd, biz_days)"
        " VALUES(?,?,?,?,?,?)",
        (s_date, e_date,
         round(final_price, 6) if final_price is not None else None,
         vol_mtd,
         round(rev_mtd, 3) if rev_mtd is not None else None,
         biz_days)
    )
    conn.commit()
    print(f"  [BULLETIN] _weekly_raw updated: {s_date} → {e_date} | "
          f"vol_MTD={vol_mtd:,.0f} t"
          + (f" | rev={rev_mtd:,.1f} 000 USD" if rev_mtd else "")
          + (f" | price_MTD={final_price:.4f} USD/kg" if final_price else "")
          + (f" | biz_days={biz_days}" if biz_days else ""))
    return 1


# ══════════════════════════════════════════════════════════════════════════════
# FILL MISSING SECEX MONTHS FROM WEEKLY AVERAGES
# ══════════════════════════════════════════════════════════════════════════════
def fill_secex_from_weekly(conn):
    """
    For months after the last official SECEX entry, estimate price_usd_kg as the
    simple average of weekly prices that fall within that month.

    Uses INSERT OR IGNORE so that when real MDIC data arrives via fetch_secex()
    (which uses INSERT OR REPLACE), the official values automatically overwrite
    these estimates.

    Returns the number of newly inserted estimated rows.
    """
    last = conn.execute(
        "SELECT year, month FROM _secex_raw ORDER BY year DESC, month DESC LIMIT 1"
    ).fetchone()
    if not last:
        return 0

    ly, lm = last
    last_date = f"{ly}-{lm:02d}-{monthrange(ly, lm)[1]:02d}"

    weekly = conn.execute(
        """
        SELECT CAST(strftime('%Y', start_date) AS INTEGER),
               CAST(strftime('%m', start_date) AS INTEGER),
               AVG(price_usd_kg)
        FROM   _weekly_raw
        WHERE  start_date > ? AND price_usd_kg IS NOT NULL
        GROUP  BY 1, 2
        ORDER  BY 1, 2
        """,
        (last_date,),
    ).fetchall()

    filled = 0
    for yr, mo, avg_p in weekly:
        if avg_p is None:
            continue
        conn.execute(
            "INSERT OR IGNORE INTO _secex_raw(year,month,rev_000usd,vol_tons,price_usd_kg)"
            " VALUES(?,?,NULL,NULL,?)",
            (yr, mo, round(avg_p, 6)),
        )
        if conn.execute("SELECT changes()").fetchone()[0]:
            print(f"  [SECEX-est] {yr}-{mo:02d}: {avg_p:.4f} USD/kg  ← weekly avg (official MDIC pending)")
            filled += 1

    conn.commit()
    return filled


# ══════════════════════════════════════════════════════════════════════════════
# COMPUTE & MATERIALISE SPREAD TABLES
# ══════════════════════════════════════════════════════════════════════════════
def _avg(conn, table, col, s, e):
    r = conn.execute(
        f"SELECT AVG({col}) FROM {table} WHERE dt >= ? AND dt <= ?", (s, e)
    ).fetchone()
    return r[0] if r and r[0] is not None else None


def materialise(conn):
    """Compute monthly and weekly spread tables from raw data."""
    now_iso = datetime.now().isoformat(timespec="seconds")

    # ── Monthly ───────────────────────────────────────────────────────────────
    raw_m = conn.execute(
        "SELECT year, month, price_usd_kg FROM _secex_raw ORDER BY year, month"
    ).fetchall()
    monthly_rows = []
    for yr, mo, p_usd in raw_m:
        s  = f"{yr}-{mo:02d}-01"
        ld = monthrange(yr, mo)[1]
        e  = f"{yr}-{mo:02d}-{ld:02d}"
        fx = _avg(conn, "_fx_raw", "fx", s, e)
        ca = _avg(conn, "_cepea_raw", "r_kg", s, e)
        brl = (p_usd * fx) if p_usd and fx else None
        sp  = (brl / ca)  if brl and ca  else None
        monthly_rows.append((
            f"{yr}-{mo:02d}", yr, mo,
            round(p_usd, 6) if p_usd else None,
            round(fx,   6) if fx else None,
            round(brl,  6) if brl else None,
            round(ca,   6) if ca else None,
            round(sp,   6) if sp else None,
            now_iso,
        ))
    conn.executemany(
        "INSERT OR REPLACE INTO monthly"
        "(period,year,month,secex_usd_kg,fx,secex_brl_kg,cepea_r_kg,spread,updated_at)"
        " VALUES(?,?,?,?,?,?,?,?,?)",
        monthly_rows
    )

    # ── Weekly ────────────────────────────────────────────────────────────────
    raw_w = conn.execute(
        "SELECT start_date, end_date, price_usd_kg, vol_tons, rev_000usd, biz_days"
        " FROM _weekly_raw ORDER BY start_date"
    ).fetchall()

    # De-accumulate MTD vol_tons AND MTD rev_000usd → incremental weekly values.
    # For price: use de-accumulated revenue/volume when available (most accurate).
    # Fallback: use stored price_usd_kg (seed data already holds incremental price).
    prev_vol_mtd = {}   # (yr, mo) -> last vol_mtd
    prev_rev_mtd = {}   # (yr, mo) -> last rev_mtd (000 USD)

    weekly_rows = []
    for s, e, p_usd, vol_mtd, rev_mtd, biz_d in raw_w:
        yr_mo = (int(s[:4]), int(s[5:7]))

        # ── De-accumulate volume ──────────────────────────────────────────────
        if vol_mtd is not None:
            prev_v   = prev_vol_mtd.get(yr_mo, 0.0)
            vol_week = vol_mtd - prev_v
            prev_vol_mtd[yr_mo] = vol_mtd
        else:
            vol_week = None

        # ── Compute incremental price from de-accumulated revenue / volume ────
        # Formula: price_USD_kg = rev_week_000USD / vol_week_tons
        #   because: (rev_000USD × 1000 USD) / (vol_tons × 1000 kg) = rev_000USD / vol_tons
        if rev_mtd is not None and vol_week is not None and vol_week > 10:
            prev_r   = prev_rev_mtd.get(yr_mo, 0.0)
            rev_week = rev_mtd - prev_r
            prev_rev_mtd[yr_mo] = rev_mtd
            price_incremental = rev_week / vol_week   # USD/kg
        elif p_usd is not None:
            # Seed rows: p_usd is already the incremental weekly price → use directly
            price_incremental = p_usd
            # Estimate rev for tracking continuity (needed when live row follows)
            if vol_week is not None and vol_week > 10:
                prev_r = prev_rev_mtd.get(yr_mo, 0.0)
                prev_rev_mtd[yr_mo] = prev_r + p_usd * vol_week
        else:
            price_incremental = None

        # ── Business days and daily average volume ────────────────────────────
        if biz_d is None:
            biz_d = _biz_days_between(
                date.fromisoformat(s), date.fromisoformat(e)
            )
        vol_daily = (vol_week / biz_d
                     if vol_week is not None and biz_d and biz_d > 0
                     else None)

        fx  = _avg(conn, "_fx_raw",    "fx",   s, e)
        ca  = _avg(conn, "_cepea_raw", "r_kg", s, e)
        brl = (price_incremental * fx) if price_incremental and fx else None
        sp  = (brl / ca)               if brl and ca              else None

        weekly_rows.append((
            s, e,
            round(price_incremental, 6) if price_incremental is not None else None,
            round(fx,                6) if fx       else None,
            round(brl,               6) if brl      else None,
            round(ca,                6) if ca       else None,
            round(sp,                6) if sp       else None,
            round(vol_week,          3) if vol_week is not None else None,
            biz_d,
            round(vol_daily,         3) if vol_daily is not None else None,
            now_iso,
        ))

    conn.executemany(
        "INSERT OR REPLACE INTO weekly"
        "(start_date,end_date,secex_usd_kg,fx,secex_brl_kg,cepea_r_kg,spread,"
        "vol_tons,biz_days,vol_tons_daily,updated_at)"
        " VALUES(?,?,?,?,?,?,?,?,?,?,?)",
        weekly_rows
    )
    conn.commit()

    nw_vol = sum(1 for r in weekly_rows if r[7] is not None)
    nm = conn.execute("SELECT COUNT(*) FROM monthly WHERE spread IS NOT NULL").fetchone()[0]
    nw = conn.execute("SELECT COUNT(*) FROM weekly  WHERE spread IS NOT NULL").fetchone()[0]
    print(f"  [DB] monthly: {len(monthly_rows)} rows ({nm} with spread)")
    print(f"  [DB] weekly:  {len(weekly_rows)} rows ({nw} with spread, {nw_vol} with volume)")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    args = sys.argv[1:]
    do_init    = "--init" in args
    cepea_idx  = next((i for i, a in enumerate(args) if a == "--cepea"), None)
    cepea_path = args[cepea_idx + 1] if cepea_idx is not None and cepea_idx + 1 < len(args) else None

    print("=" * 60)
    print(f"  Brazil Beef Spread Extractor — {date.today().isoformat()}")
    print(f"  Mode: {'INIT (full seed)' if do_init else 'INCREMENTAL UPDATE'}")
    print("=" * 60)

    conn = sqlite3.connect(DB_PATH)
    init_db(conn)

    # Weekly raw always seeded (INSERT OR REPLACE → idempotent)
    print("\n[1] Seeding weekly raw data …")
    seed_weekly_raw(conn)

    if do_init:
        # Full SECEX history
        print("\n[2] Fetching SECEX monthly (all years) …")
        fetch_secex(conn)

        # CEPEA from local XLS
        print("\n[3] Loading CEPEA from local XLS …")
        if cepea_path:
            load_cepea_xls(conn, cepea_path)
        else:
            candidates = (
                list(Path(DB_PATH.parent).glob("CEPEA*.xls*")) +
                list(Path(DB_PATH.parent).glob("cepea*.xls*")) +
                list(Path(DB_PATH.parent.parent).glob("CEPEA*.xls*"))
            )
            if candidates:
                load_cepea_xls(conn, candidates[0])
            else:
                print("  [CEPEA] No XLS file found.")
                print("  Run with:  python extractor_bz.py --init --cepea /path/to/CEPEA.xls")
                print("  Download:  https://www.cepea.esalq.usp.br/br/indicador/boi-gordo.aspx")

        print("\n[4] Fetching BCB PTAX FX …")
        fetch_fx(conn)

    else:
        # Incremental: current + prior year SECEX only
        print("\n[2] Fetching SECEX monthly (recent years) …")
        yr = datetime.now().year
        fetch_secex(conn, years=[yr - 1, yr])

        print("\n[3] Fetching BCB PTAX FX …")
        fetch_fx(conn)

        # Fetch latest CEPEA data from web (incremental only — init uses local XLS)
        print("\n[3b] Fetching CEPEA Boi Gordo from web …")
        fetch_cepea_web(conn)

    # Weekly bulletin: fetch latest price + MTD volume (both init and incremental)
    print("\n[→] Fetching SECEX weekly bulletin (price + MTD volume) …")
    fetch_weekly_bulletin(conn)

    print("\n[→] Computing spread tables …")
    materialise(conn)

    conn.close()
    print("\n" + "=" * 60)
    print("  Done. beef_bz.db updated.")
    print("=" * 60)


if __name__ == "__main__":
    main()
